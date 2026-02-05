"""
Control de Lotes - Versión Flet (Android/Desktop/Web)
Adaptado de lotes_gui.py (Tkinter) para funcionar en múltiples plataformas.
"""

import flet as ft
import asyncio
import csv
import os
import sys
import base64
import json
from datetime import datetime
import requests
import shutil
import glob
import hashlib
import time

# Importar fpdf2 para exportar PDF (opcional)
try:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False
    XPos = None
    YPos = None

# Importar openpyxl para exportar Excel (opcional)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ========== CONFIGURACIÓN ==========

if getattr(sys, 'frozen', False):
    BASE_PATH = os.path.dirname(sys.executable)
else:
    BASE_PATH = os.path.dirname(os.path.abspath(__file__))

# En Android, el directorio de trabajo suele apuntar al área de datos de la app.
# Preferimos usar el cwd en ese caso para almacenar y buscar archivos.
if hasattr(sys, 'getandroidapilevel'):
    try:
        BASE_PATH = os.getcwd()
    except Exception:
        pass

CONFIG_FILE = os.path.join(BASE_PATH, "github_config.txt")
LOTES_CSV = os.path.join(BASE_PATH, "lotes_template.csv")
# Archivo local de trabajo si el usuario 'borra' datos en la UI; preservamos el original
LOTES_WORKING = os.path.join(BASE_PATH, "lotes_local.csv")
REGISTROS_DIR = os.path.join(BASE_PATH, "registros")

# Debug: mostrar rutas usadas (útil en Android para detectar problemas de path)
try:
    print(f"[PATHS] BASE_PATH={BASE_PATH} CONFIG_FILE={CONFIG_FILE} LOTES_CSV={LOTES_CSV}")
except Exception:
    pass
# Sentinel file to disable automatic restore after user clears local data
NO_AUTO_RESTORE_FILE = os.path.join(BASE_PATH, ".no_auto_restore")
# Timestamp for last config clear to avoid races when reading SharedPreferences
CONFIG_LAST_CLEARED = 0.0
# Flag to indicate local data was cleared; used to avoid accidental uploads
LOCAL_DATA_CLEARED = False

VERSION = '1.0.5'
BRANCH = ['FSM', 'SMB', 'RP']
STAGES = ['CLONADO', 'VEG. TEMPRANO', 'VEG. TARDIO', 'FLORACIÓN', 'TRANSICIÓN', 'SECADO', 'PT']
LOCATIONS = ['PT', 'CUARTO 1', 'CUARTO 2', 'CUARTO 3', 'CUARTO 4', 'VEGETATIVO', 'ENFERMERÍA', 'MADRES']
VARIETIES = [
    'Ak-47', 'Apple Fritter', 'Banana Latte', 'Blackberry Honey', 'Desconocida',
    'Gran Jefa', 'HG23 (Michael Jordan)', 'Kandy Kush', 'King Kush Breath',
    'Kosher Kush', 'Mozzerella', 'Orangel', 'Purple Diesel', 'ReCon',
    'Red Red Wine', 'Runtz', 'Sugar Cane', 'Wedding Cake', 'Zallah Bread',
]

# Variables globales
GITHUB_REPO = ""
GITHUB_TOKEN = ""
GITHUB_FILE_PATH = "lotes_template.csv"
GITHUB_BRANCH = "main"
CURRENT_USER = ""  # Usuario actual de la app


def normalizar_nombre(nombre: str) -> str:
    """Normaliza un nombre: 'eDuaRdO' -> 'Eduardo', 'JUAN PABLO' -> 'Juan Pablo'"""
    if not nombre:
        return ""
    return ' '.join(word.capitalize() for word in nombre.strip().split())


def get_config_path():
    """Obtiene la ruta del archivo de configuración según la plataforma (solo desktop)."""
    return os.path.join(BASE_PATH, "lotes_config.json")


def encontrar_ruta_config():
    """Busca el archivo de configuración en varias ubicaciones posibles y retorna la ruta si existe."""
    candidates = [
        get_config_path(),
        os.path.join(os.getcwd(), "lotes_config.json"),
        os.path.join(BASE_PATH, "_lotes_config.json"),
    ]
    # Buscar cualquier archivo que contenga 'lotes_config' en BASE_PATH
    for p in glob.glob(os.path.join(BASE_PATH, "*lotes_config*.json")):
        candidates.append(p)

    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def user_has_config():
    """Retorna True si el usuario tiene repo, token y usuario configurados."""
    return bool(GITHUB_REPO and GITHUB_TOKEN and CURRENT_USER)


# Reactivation UI removed: reactivation must be performed by configuring GitHub
# and manually removing the sentinel '.no_auto_restore' outside the app. The
# explicit button and helper were removed to avoid accidental reactivation.


def cargar_config_desde_storage(page=None):
    """Carga la configuración desde archivo JSON (desktop) o client_storage (Android)."""
    global GITHUB_REPO, GITHUB_TOKEN, CURRENT_USER
    if hasattr(sys, 'getandroidapilevel') and page is not None:
        # Android: usar SharedPreferences async
        from flet import SharedPreferences
        async def get_config():
            try:
                prefs = SharedPreferences()
                config_str = await prefs.get("lotes_config")
                if config_str:
                    config = json.loads(config_str)
                    repo = config.get("github_repo", "")
                    token = config.get("github_token", "")
                    globals()["CURRENT_USER"] = config.get("current_user", "")
                    if repo and token and "/" in repo:
                        globals()["GITHUB_REPO"] = repo
                        globals()["GITHUB_TOKEN"] = token
                        return True, "Config cargada"
                return False, "Configura GitHub en ⚙️"
            except Exception as e:
                return False, f"Error config: {e}"
        return get_config  # Devuelve la función async para ser llamada con await o create_task
    else:
        # Desktop: archivo
        config_path = encontrar_ruta_config()
        
        if config_path and os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                repo = config.get("github_repo", "")
                token = config.get("github_token", "")
                user = config.get("current_user", "")
                
                # Asignar siempre los valores si existen
                GITHUB_REPO = repo
                GITHUB_TOKEN = token
                CURRENT_USER = user
                # Además, si se leyó desde una ruta distinta, reescribir config en la ruta esperada
                expected = get_config_path()
                if os.path.abspath(config_path) != os.path.abspath(expected):
                    try:
                        with open(expected, 'w', encoding='utf-8') as f:
                            json.dump(config, f, ensure_ascii=False, indent=2)
                    except Exception as e:
                        # No crítico si no podemos copiar el archivo; continuar con lo que tenemos
                        pass

                if repo and token and "/" in repo:
                    return True, "Config cargada"
                else:
                    
                    # Retornar True si al menos hay usuario (queremos mostrarlo)
                    if user:
                        return True, "Usuario cargado"
                    return False, "Config incompleta, revisa ⚙️"
            except Exception as e:
                
                return False, f"Error config: {e}"
        
        return False, "Configura GitHub en ⚙️"


def guardar_config_en_storage(page, repo, token, user=None):
    """Guarda la configuración en archivo JSON (desktop) o client_storage (Android)."""
    global GITHUB_REPO, GITHUB_TOKEN, CURRENT_USER
    config = {}
    if hasattr(sys, 'getandroidapilevel') and page is not None:
        # Android: usar SharedPreferences async
        try:
            from flet import SharedPreferences
            async def set_config():
                prefs = SharedPreferences()
                config["github_repo"] = repo
                config["github_token"] = token
                if user:
                    config["current_user"] = user
                await prefs.set("lotes_config", json.dumps(config, ensure_ascii=False))
                globals()["GITHUB_REPO"] = repo
                globals()["GITHUB_TOKEN"] = token
                if user:
                    globals()["CURRENT_USER"] = user
            # Usar create_task en Android para no bloquear el event loop
            try:
                asyncio.create_task(set_config())
            except RuntimeError:
                # Si no hay loop activo, ejecutar en hilo
                asyncio.run(set_config())
        except Exception as e:
            print(f"Error guardando config: {e}")
            # Fallback: si SharedPreferences falla en Android, asignar globals para permitir
            # que la UI continúe (persistencia puede no estar garantizada sin SharedPreferences).
            try:
                globals()["GITHUB_REPO"] = repo
                globals()["GITHUB_TOKEN"] = token
                if user:
                    globals()["CURRENT_USER"] = user
            except Exception as _:
                pass
    else:
        # Desktop: archivo
        config_path = get_config_path()
        existing_config = {}
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    existing_config = json.load(f)
            except:
                pass
        config = existing_config
        config["github_repo"] = repo
        config["github_token"] = token
        if user:
            config["current_user"] = user
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        GITHUB_REPO = repo
        GITHUB_TOKEN = token
        if user:
            CURRENT_USER = user


def guardar_usuario(nombre: str):
    """Guarda solo el usuario en la configuración."""
    global CURRENT_USER

    nombre_normalizado = normalizar_nombre(nombre)

    # Preferir usar la función general para persistir en el medio correcto (Android vs Desktop)
    try:
        # Pasar None como page ya que guardar_config_en_storage maneja Android internamente
        guardar_config_en_storage(None, GITHUB_REPO or "", GITHUB_TOKEN or "", user=nombre_normalizado)
    except Exception:
        # Fallback: escribir localmente
        config_path = get_config_path()
        config = {}
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            except:
                pass
        config["current_user"] = nombre_normalizado
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2)

    CURRENT_USER = nombre_normalizado
    return nombre_normalizado


def descargar_csv_github():
    """Descarga el CSV desde GitHub y guarda como local si no hay conflicto.
    Devuelve (success, msg)."""
    print("[NETWORK] descargar_csv_github: inicio")
    ok, msg, remote_content, remote_hash = get_remote_csv_content()
    if not ok:
        print(f"[NETWORK] descargar_csv_github: no ok -> {msg}")
        return False, msg

    # Leer estado local/meta
    meta = load_local_meta()
    local_content = ''
    if os.path.exists(LOTES_CSV):
        try:
            with open(LOTES_CSV, 'r', encoding='utf-8') as f:
                local_content = f.read()
        except Exception:
            local_content = ''
    local_hash = compute_hash(local_content) if local_content else ''

    # Si local está vacío o igual al remoto (no cambios), escribir remoto
    if not local_content or local_hash == remote_hash:
        # Crear backup del local si existe
        if local_content:
            b = crear_backup()
            if b:
                print(f"[NETWORK] descargar_csv_github: backup local creado {b}")
        try:
            with open(LOTES_CSV, 'w', encoding='utf-8') as f:
                f.write(remote_content)
            fix_csv_structure()
            # Actualizar meta
            meta['local_hash'] = remote_hash
            meta['remote_hash'] = remote_hash
            save_local_meta(meta)
            return True, 'Conectado'
        except Exception as e:
            print(f"[NETWORK] error escribiendo local: {e}")
            return False, f'Error escritura: {e}'

    # Si hay diferencias y local cambió desde el último remoto conocido -> conflicto
    if meta.get('local_hash') and meta.get('local_hash') != remote_hash and local_hash != meta.get('remote_hash'):
        # Guardar ambos en registros para revisión manual y no sobrescribir
        b = crear_backup()
        rb = save_remote_backup(remote_content)
        print(f"[NETWORK] conflicto remoto/local: backup_local={b} backup_remote={rb}")
        return False, 'Conflicto local/remoto, backups guardados'

    # Si local no fue modificado desde último remote conocido, entonces remote es la fuente -> sobrescribir
    try:
        with open(LOTES_CSV, 'w', encoding='utf-8') as f:
            f.write(remote_content)
        fix_csv_structure()
        meta['local_hash'] = remote_hash
        meta['remote_hash'] = remote_hash
        save_local_meta(meta)
        return True, 'Conectado'
    except Exception as e:
        print(f"[NETWORK] error escribiendo (2): {e}")
        return False, f'Error escritura: {e}'


def subir_csv_github(force: bool = False):
    """Sube el CSV a GitHub. Devuelve (success, msg). Maneja conflictos basados en meta local/remote."""
    print("[NETWORK] subir_csv_github: inicio")
    # Validaciones: token, repo y usuario
    if not GITHUB_TOKEN:
        print("[NETWORK] subir_csv_github: sin token")
        return False, 'Sin token'
    if not GITHUB_REPO or "/" not in GITHUB_REPO:
        print("[NETWORK] subir_csv_github: repo no configurado")
        return False, 'Repo no configurado'
    if not CURRENT_USER:
        print("[NETWORK] subir_csv_github: falta usuario")
        return False, 'Falta usuario configurado (⚙️ Usuario)'

    url = f'https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE_PATH}'
    headers = {
        'Authorization': f'token {GITHUB_TOKEN}',
        'Accept': 'application/vnd.github.v3+json'
    }

    # Evitar subir si el usuario borró datos locales y no reactivó manualmente
    try:
        if globals().get('LOCAL_DATA_CLEARED'):
            print("[NETWORK] subir_csv_github: upload bloqueado porque se borraron datos locales recientemente (requiere reactivar subidas)")
            return False, 'Subidas bloqueadas tras borrar datos locales. Reactiva subidas en Config para continuar.'
    except Exception:
        pass

    # Leer meta y local
    meta = load_local_meta()
    try:
        with open(LOTES_CSV, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"[NETWORK] subir_csv_github: error leyendo local: {e}")
        return False, 'Error lectura local'
    # Comprobar si el CSV local tiene datos útiles (más allá del encabezado)
    try:
        import io
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        data_rows = [r for r in rows[1:] if any((c or '').strip() for c in r)] if len(rows) > 1 else []
        if not data_rows and not force:
            print("[NETWORK] subir_csv_github: local vacío o sólo cabecera, abortando")
            return False, 'Local vacío o sólo cabecera, usa force=True para forzar subida'
    except Exception:
        # Si falla al analizar, continuar con hash calculado
        pass

    local_hash = compute_hash(content)

    # Consultar remoto breve para detectar cambios
    try:
        resp = requests.get(url, headers=headers, params={'ref': GITHUB_BRANCH}, timeout=5)
        remote_content = ''
        if resp.status_code == 200:
            remote_content = base64.b64decode(resp.json().get('content', '')).decode('utf-8')
            remote_hash = compute_hash(remote_content)
        elif resp.status_code == 404:
            # No existe el archivo remoto
            remote_hash = ''
            remote_content = ''
            if not force:
                return False, f'Archivo {GITHUB_FILE_PATH} no encontrado', '', ''
            # Si force==True, permitimos crear el archivo más abajo (sólo si hay datos locales)
        else:
            remote_hash = ''
    except Exception:
        remote_hash = ''

    # Si hay conflicto (remote cambió desde último conocido y local también cambió) -> abortar
    if (remote_hash and meta.get('remote_hash') and remote_hash != meta.get('remote_hash') and meta.get('local_hash') and local_hash != meta.get('local_hash')) and not force:
        print("[NETWORK] subir_csv_github: conflicto detectado, abortando para evitar sobrescribir")
        # Guardar remote para revisión
        if remote_content:
            rb = save_remote_backup(remote_content)
            print(f"[NETWORK] subir_csv_github: backup remoto guardado {rb}")
        return False, 'Conflicto remoto detectado'

    # Proceder a subir
    try:
        encoded_content = base64.b64encode(content.encode('utf-8')).decode('utf-8')
        fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M")
        commit_msg = f"Actualización {fecha_hora} {CURRENT_USER}"
        data = {'message': commit_msg, 'content': encoded_content, 'branch': GITHUB_BRANCH}
        if resp.status_code == 200:
            data['sha'] = resp.json().get('sha', '')

        # Si resp.status_code == 404 y force is True, permitimos crear el archivo solo si hay datos locales
        if resp.status_code == 404:
            try:
                import io
                reader = csv.reader(io.StringIO(content))
                rows = list(reader)
                data_rows = [r for r in rows[1:] if any((c or '').strip() for c in r)] if len(rows) > 1 else []
                if not data_rows:
                    print("[NETWORK] subir_csv_github: no se crea archivo remoto vacío")
                    return False, 'No se crea archivo remoto vacío'
            except Exception:
                # Si no podemos analizar, ser conservadores: no crear
                return False, 'No se puede crear remoto sin datos'

        # Crear backup remoto previo (por seguridad) si existe contenido remoto
        try:
            if remote_content:
                rb_prev = save_remote_backup(remote_content)
                print(f"[NETWORK] subir_csv_github: backup remoto previo creado {rb_prev}")
        except Exception:
            pass

        response = requests.put(url, headers=headers, json=data, timeout=10)
        print(f"[NETWORK] subir_csv_github: put status={response.status_code}")
        if response.status_code in [200, 201]:
            # Crear backup local DESPUÉS de sincronizar exitosamente
            crear_backup()
            # Actualizar meta
            meta['local_hash'] = local_hash
            meta['remote_hash'] = local_hash
            save_local_meta(meta)
            return True, 'Sincronizado'
        else:
            return False, f'Error {response.status_code}'
    except Exception as e:
        print(f"[NETWORK] subir_csv_github: exception {e}")
        return False, f'Error: {str(e)[:50]}'


def leer_csv():
    """Lee lotes del CSV."""
    # Si el usuario marcó borrado local, preferimos el archivo de trabajo o devolver vacío sin tocar el original
    try:
        if globals().get('LOCAL_DATA_CLEARED'):
            if os.path.exists(LOTES_WORKING):
                csv_path = LOTES_WORKING
            else:
                return []
        else:
            csv_path = LOTES_CSV
    except Exception:
        csv_path = LOTES_CSV

    # Fallback: si no existe en csv_path (p.e. en Android), buscar en cwd y otras rutas comunes
    if not os.path.exists(csv_path):
        alt = os.path.join(os.getcwd(), os.path.basename(csv_path))
        if os.path.exists(alt):
            csv_path = alt
            try:
                print(f"[PATHS] usar csv alternativo: {csv_path}")
            except Exception:
                pass
        else:
            # Probar registros por si hay un backup reciente
            backups = glob.glob(os.path.join(REGISTROS_DIR, "lotes_template_*.csv"))
            if backups:
                # No restauramos automáticamente aquí, pero podemos preferir el backup como última fuente
                try:
                    csv_path = backups[-1]
                    print(f"[PATHS] usar backup como fuente: {csv_path}")
                except Exception:
                    pass
    if not os.path.exists(csv_path):
        return []
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            lotes_final = []
            for row in reader:
                variedades = []
                for i in range(1, 21):
                    v = row.get(f'Variedad_{i}', '').strip()
                    c = row.get(f'Cantidad_{i}', '').strip()
                    if v:
                        try:
                            c = int(c)
                        except:
                            c = 0
                        variedades.append({'name': v, 'count': c})
                row['Variedades'] = variedades
                lotes_final.append(row)
            return lotes_final
    except Exception as e:
        try:
            print(f"[READ] error leyendo CSV {csv_path}: {e}")
        except Exception:
            pass
        return []


def guardar_csv(lotes):
    """Guarda lotes en el CSV. Si el usuario marcó borrado local, escribimos en el archivo de trabajo para preservar el original."""
    target = LOTES_WORKING if globals().get('LOCAL_DATA_CLEARED') else LOTES_CSV
    try:
        with open(target, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['ID', 'Branch', 'LoteNum', 'Stage', 'Location', 'Semana', 
                         'DateCreated', 'ÚltimaActualización', 'Notes']
            for i in range(1, 21):
                fieldnames.extend([f'Variedad_{i}', f'Cantidad_{i}'])
            
            writer = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
            writer.writeheader()
            
            for row in lotes:
                if 'ÚltimaActualización' not in row:
                    row['ÚltimaActualización'] = ''
                if 'Variedades' in row:
                    variedades = row['Variedades']
                    for i in range(1, 21):
                        row[f'Variedad_{i}'] = ''
                        row[f'Cantidad_{i}'] = ''
                    for i, v in enumerate(variedades[:20], start=1):
                        row[f'Variedad_{i}'] = v.get('name', '')
                        row[f'Cantidad_{i}'] = str(v.get('count', 0))
                    del row['Variedades']
                writer.writerow(row)
        # Actualizar hash local en meta
        try:
            with open(target, 'r', encoding='utf-8') as f:
                content = f.read()
            meta = load_local_meta()
            meta['local_hash'] = compute_hash(content)
            save_local_meta(meta)
        except Exception:
            pass
        return True
    except Exception:
        return False


def ensure_registros_dir():
    os.makedirs(REGISTROS_DIR, exist_ok=True)


def crear_backup():
    # Preferir respaldar el archivo de trabajo si existe (el que refleja el estado local activo)
    target = LOTES_WORKING if os.path.exists(LOTES_WORKING) else LOTES_CSV
    if not os.path.exists(target):
        return None
    ensure_registros_dir()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dest = os.path.join(REGISTROS_DIR, f"lotes_template_{timestamp}.csv")
    try:
        shutil.copy2(target, dest)
        return dest
    except Exception:
        return None


# --- Meta and hashing helpers for conflict detection ---
def get_local_meta_path():
    return os.path.join(BASE_PATH, 'lotes_local_meta.json')


def load_local_meta():
    path = get_local_meta_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def save_local_meta(meta: dict):
    path = get_local_meta_path()
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def compute_hash(text: str) -> str:
    try:
        return hashlib.sha256(text.encode('utf-8')).hexdigest()
    except Exception:
        return ''


def save_remote_backup(content: str):
    """Guarda el contenido remoto como backup en registros para revisión manual."""
    ensure_registros_dir()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dest = os.path.join(REGISTROS_DIR, f"remote_lotes_{timestamp}.csv")
    try:
        with open(dest, 'w', encoding='utf-8') as f:
            f.write(content)
        return dest
    except Exception:
        return None


def get_remote_csv_content():
    """Obtiene el contenido remoto (sin escribir localmente). Devuelve (success, msg, content, hash)
    Mejora: prueba ramas alternativas (p.ej. 'main' y 'master') y verifica existencia del repo para mensajes más claros."""
    if not GITHUB_TOKEN:
        return False, 'Sin token configurado', '', ''

    headers = {
        'Authorization': f'token {GITHUB_TOKEN}',
        'Accept': 'application/vnd.github.v3+json'
    }

    tried_branches = []
    branches_to_try = []
    # Priorizar la rama configurada, luego 'main' y 'master'
    if GITHUB_BRANCH:
        branches_to_try.append(GITHUB_BRANCH)
    for b in ('main', 'master'):
        if b not in branches_to_try:
            branches_to_try.append(b)

    url_base = f'https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE_PATH}'

    try:
        for br in branches_to_try:
            tried_branches.append(br)
            try:
                resp = requests.get(url_base, headers=headers, params={'ref': br}, timeout=6)
            except requests.exceptions.Timeout:
                return False, 'Timeout', '', ''

            if resp.status_code == 200:
                content = base64.b64decode(resp.json().get('content', '')).decode('utf-8')
                h = compute_hash(content)
                # Actualizar branch para reflejar la rama efectiva
                globals()['GITHUB_BRANCH'] = br
                return True, 'OK', content, h
            elif resp.status_code == 401:
                return False, 'Token inválido o sin permisos', '', ''
            elif resp.status_code == 404:
                # intentar siguiente rama
                continue
            else:
                return False, f'Error HTTP {resp.status_code}', '', ''

        # Si llegamos aquí, ninguna rama tuvo el archivo: verificar si el repo existe / hay acceso
        repo_url = f'https://api.github.com/repos/{GITHUB_REPO}'
        try:
            r = requests.get(repo_url, headers=headers, timeout=5)
            if r.status_code == 200:
                return False, f'Archivo {GITHUB_FILE_PATH} no encontrado (probadas ramas: {",".join(tried_branches)})', '', ''
            elif r.status_code == 401:
                return False, 'Token inválido o sin permisos', '', ''
            elif r.status_code == 404:
                return False, 'Repositorio no encontrado o sin acceso (verifica owner/repo)', '', ''
            else:
                return False, f'Error HTTP {r.status_code}', '', ''
        except requests.exceptions.Timeout:
            return False, 'Timeout comprobando repo', '', ''
        except Exception as e:
            return False, f'Error: {str(e)[:50]}', '', ''

    except Exception as e:
        return False, f'Error: {str(e)[:50]}', '', ''


def restore_latest_backup():
    """Restaura el backup más reciente desde /registros al archivo local."""
    ensure_registros_dir()
    files = glob.glob(os.path.join(REGISTROS_DIR, "lotes_template_*.csv"))
    if not files:
        return False, 'No hay backups disponibles'
    files.sort()
    latest = files[-1]
    try:
        shutil.copy2(latest, LOTES_CSV)
        fix_csv_structure()
        return True, f'Restaurado backup {os.path.basename(latest)}'
    except Exception as e:
        return False, f'Error restaurando backup: {e}'


# reactivate_uploads removed: reactivation via UI is intentionally disabled to avoid accidental remote changes.
# To reactivate uploads the sentinel file `.no_auto_restore` must be removed manually by the user.


def restore_remote_from_content(content: str):
    """Crea o actualiza el archivo remoto en GitHub usando el contenido proporcionado.
    Devuelve (success, msg)."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        return False, 'Token o repo no configurado'
    url = f'https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE_PATH}'
    headers = {
        'Authorization': f'token {GITHUB_TOKEN}',
        'Accept': 'application/vnd.github.v3+json'
    }
    try:
        # Verificar existencia para obtener SHA
        try:
            resp = requests.get(url, headers=headers, params={'ref': GITHUB_BRANCH}, timeout=5)
        except Exception as ex:
            return False, f'Error comprobando remoto: {ex}'
        sha = None
        if resp.status_code == 200:
            sha = resp.json().get('sha', '')
        # Preparar payload
        encoded = base64.b64encode(content.encode('utf-8')).decode('utf-8')
        fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M")
        data = {'message': f"Restauración {fecha_hora}", 'content': encoded, 'branch': GITHUB_BRANCH}
        if sha:
            data['sha'] = sha
        # Ejecutar PUT
        put = requests.put(url, headers=headers, json=data, timeout=15)
        if put.status_code in (200, 201):
            return True, 'Remoto restaurado'
        else:
            return False, f'Error {put.status_code} al restaurar remoto'
    except Exception as ex:
        return False, f'Error: {ex}'


def subir_csv_github_from_content(content: str, allow_create: bool = False):
    """Helper que sube contenido dado al archivo remoto. allow_create permite crear el archivo si no existe."""
    # Similar a subir_csv_github pero con contenido en memoria
    if not GITHUB_TOKEN:
        return False, 'Sin token'
    if not GITHUB_REPO or "/" not in GITHUB_REPO:
        return False, 'Repo no configurado'
    if not CURRENT_USER:
        return False, 'Falta usuario configurado (⚙️ Usuario)'

    url = f'https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE_PATH}'
    headers = {
        'Authorization': f'token {GITHUB_TOKEN}',
        'Accept': 'application/vnd.github.v3+json'
    }
    # Evitar subir si el usuario borró datos locales y no reactivó manualmente
    try:
        if globals().get('LOCAL_DATA_CLEARED'):
            print("[NETWORK] subir_csv_github_from_content: upload bloqueado porque se borraron datos locales recientemente (requiere reactivar subidas)")
            return False, 'Subidas bloqueadas tras borrar datos locales. Reactiva subidas en Config para continuar.'
    except Exception:
        pass
    # Comprobar si existe remotamente
    try:
        resp = requests.get(url, headers=headers, params={'ref': GITHUB_BRANCH}, timeout=6)
    except Exception as ex:
        return False, f'Error comprobando remoto: {ex}'
    if resp.status_code == 404 and not allow_create:
        return False, 'Archivo remoto no encontrado; no se crea sin permiso explícito'

    # Verificar que haya datos útiles
    try:
        import io
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        data_rows = [r for r in rows[1:] if any((c or '').strip() for c in r)] if len(rows) > 1 else []
        if not data_rows:
            return False, 'Contenido vacío: no se sube'
    except Exception:
        return False, 'Contenido no válido'

    # Construir payload y PUT (manejo de sha si existe)
    try:
        encoded_content = base64.b64encode(content.encode('utf-8')).decode('utf-8')
        data = {'message': f'Restore {datetime.now().strftime("%Y-%m-%d %H:%M")}', 'content': encoded_content, 'branch': GITHUB_BRANCH}
        if resp.status_code == 200:
            data['sha'] = resp.json().get('sha', '')
        # Guardar backup remoto previo si existe
        try:
            if resp.status_code == 200:
                remote_content = base64.b64decode(resp.json().get('content', '')).decode('utf-8')
                rb_prev = save_remote_backup(remote_content)
                print(f"[NETWORK] subir_csv_github_from_content: backup remoto previo creado {rb_prev}")
        except Exception:
            pass
        put = requests.put(url, headers=headers, json=data, timeout=15)
        if put.status_code in (200,201):
            return True, 'Remoto restaurado'
        else:
            return False, f'Error {put.status_code} al subir'
    except Exception as ex:
        return False, f'Error: {ex}'


def fix_csv_structure():
    """Normaliza la estructura del CSV local: asegura columnas correctas y mueve fechas mal colocadas."""
    try:
        lotes = leer_csv()
        # Corregir caso donde la fecha quedó en la columna 'Semana'
        for row in lotes:
            sem_val = row.get('Semana', '')
            if sem_val and isinstance(sem_val, str) and '-' in sem_val and sem_val.strip()[0].isdigit():
                row['DateCreated'] = sem_val.strip()
                row['Semana'] = ''
        guardar_csv(lotes)
    except Exception:
        pass


def startup_restore():
    """Al iniciar, descargar desde GitHub (referencia). Solo usar backup si no hay conexión.
    Evita restaurar backup automáticamente en caso de conflicto remoto/local; en ese caso reporta y deja para resolución manual."""
    # Si el usuario borró datos manualmente recientemente, evitar restauración automática
    try:
        if os.path.exists(NO_AUTO_RESTORE_FILE):
            return False, 'Auto-restore deshabilitado por acción del usuario'
    except Exception:
        pass

    success, msg = descargar_csv_github()
    if success:
        fix_csv_structure()
        return True, 'Sincronizado con GitHub'
    else:
        # Si la falla fue un conflicto, no hacemos restauración automática
        if isinstance(msg, str) and 'Conflicto' in msg:
            return False, msg
        # Si no hay conexión o error, intentar usar backup local
        ok, info = restore_latest_backup()
        if ok:
            return True, f'Offline: {info}'
        else:
            return False, msg


def find_lote_by_id(lote_id, lotes=None):
    """Busca un lote por su ID, considerando ubicación si está presente."""
    if lotes is None:
        lotes = leer_csv()
    
    sel = lote_id.strip()
    location_filter = None
    
    # Extraer ubicación si está en formato "L1-FSM (CUARTO 1)"
    if '(' in sel and sel.endswith(')'):
        parts = sel.rsplit('(', 1)
        sel = parts[0].strip()
        location_filter = parts[1].rstrip(')').strip()
    
    if '|' in sel:
        sel = sel.split('|', 1)[0].strip()
    
    for idx, lote in enumerate(lotes):
        calc_id = f"L{lote.get('LoteNum')}-{lote.get('Branch')}"
        if sel == calc_id or sel == lote.get('ID'):
            # Si hay filtro de ubicación, verificar que coincida
            if location_filter:
                if lote.get('Location', '') == location_filter:
                    return idx, lote
            else:
                return idx, lote
    return None, None


def get_lote_ids_sorted():
    """Retorna lista de IDs de lotes ordenados."""
    lotes = leer_csv()
    
    def lote_key(lote):
        try:
            return (lote.get('Branch', ''), int(lote.get('LoteNum', 0)))
        except:
            return (lote.get('Branch', ''), 0)
    
    lotes_sorted = sorted(lotes, key=lote_key)
    
    counts = {}
    for lote in lotes_sorted:
        cid = f"L{lote.get('LoteNum')}-{lote.get('Branch')}"
        counts[cid] = counts.get(cid, 0) + 1
    
    ids = []
    for lote in lotes_sorted:
        cid = f"L{lote.get('LoteNum')}-{lote.get('Branch')}"
        if counts.get(cid, 0) > 1:
            label = f"{cid} ({lote.get('Location', '')})"
        else:
            label = cid
        ids.append(label)
    return ids


# ========== APLICACIÓN FLET ==========

def main(page: ft.Page):
    page.title = "Control de Lotes"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.padding = 10
    # Inicialización asíncrona para Android y Desktop
    
    async def init_config():
        global GITHUB_REPO, GITHUB_TOKEN, CURRENT_USER

        config_ok, config_msg = False, ""
        if hasattr(sys, 'getandroidapilevel'):
            
            get_config = cargar_config_desde_storage(page)
            if get_config:
                config_ok, config_msg = await get_config()
        else:
            
            try:
                config_ok, config_msg = cargar_config_desde_storage(page)
                
                
            except Exception as e:
                print(f"Error en cargar_config_desde_storage: {e}")
                page.add(ft.Text(f"Error en inicialización: {e}", color=ft.Colors.RED, size=16))
                page.update()
                return
        # Actualizar sólo el mensaje de estado en la pestaña config (TextFields actualizan al abrir la pestaña)
        config_status.value = config_msg
        config_status.color = ft.Colors.GREEN if config_ok else ft.Colors.RED
        # Usar la verificación detallada para el texto y color de la barra de estado
        try:
            check_and_update_connection_status()
        except Exception:
            # Fallback: establecer mensaje simple
            if status_text and status_text.current:
                status_text.current.value = config_msg or "No conectado"
                if connection_status and connection_status.current:
                    connection_status.current.bgcolor = ft.Colors.RED_400
        page.update()
        

    # Lanzar la inicialización asíncrona al cargar la página
    def on_page_load(e):
        async def startup():
            # Mostrar UI inmediatamente para evitar pantallas intermedias y que el usuario vea algo rápido
            try:
                content_area.content = ft.Container(tab_crear, padding=15)
                if status_text and status_text.current:
                    status_text.current.value = "Cargando..."
                page.update()
            except Exception:
                pass

            # Lanzar inicialización en background (no bloqueante)
            try:
                asyncio.create_task(init_config())
            except Exception as ex:
                print(f"[STARTUP] no se pudo lanzar init_config en background: {ex}")

            # Deferir restauración/descarga a tarea en background para que no bloquee la UI
            async def background_restore():
                await asyncio.sleep(0.4)
                try:
                    if status_text and status_text.current:
                        status_text.current.value = "Restaurando datos..."
                        page.update()

                    # Intentar preferir remoto (ambos: Android y Desktop) pero sin bloquear la UI
                    try:
                        ok, info = await asyncio.wait_for(asyncio.to_thread(startup_restore), timeout=8)
                        if ok:
                            show_snackbar(info)
                            try:
                                update_status(True, info)
                            except Exception:
                                pass
                            # Refrescar datos y UI ahora que se restauró remoto
                            try:
                                refresh_lotes_list_radios()
                                refresh_edit_lotes_popup()
                            except Exception:
                                pass
                        else:
                            # Si el usuario deshabilitó auto-restore, no intentar restaurar backups
                            if isinstance(info, str) and info.startswith('Auto-restore'):
                                show_snackbar(info)
                                try:
                                    update_status(False, info)
                                except Exception:
                                    pass
                            # Si hubo un conflicto, startup_restore retornó mensaje con 'Conflicto'
                            elif isinstance(info, str) and 'Conflicto' in info:
                                # Mostrar diálogo para que el usuario elija restaurar remoto o mantener local
                                def cerrar_conf(e):
                                    dlg_conf.open = False
                                    page.update()

                                def mantener_local(e):
                                    dlg_conf.open = False
                                    page.update()
                                    show_snackbar('Manteniendo datos locales')
                                    try:
                                        update_status(False, 'Conflicto: revisar backups')
                                    except Exception:
                                        pass

                                dlg_conf = ft.AlertDialog(
                                    modal=True,
                                    title=ft.Text('Conflicto de inicio'),
                                    content=ft.Text('Se detectó una diferencia entre remoto y local al iniciar. Se mantendrán los datos locales para evitar sobrescribir remoto.'),
                                    actions=[
                                        ft.TextButton('OK', on_click=cerrar_conf),
                                        ft.TextButton('Mantener local', on_click=mantener_local),
                                    ],
                                    actions_alignment=ft.MainAxisAlignment.END,
                                )
                                page.overlay.append(dlg_conf)
                                dlg_conf.open = True
                                page.update()
                            else:
                                # No hay conexión o error; intentar fallback a backup
                                show_snackbar(info, error=True)
                                try:
                                    ok2, info2 = await asyncio.to_thread(restore_latest_backup)
                                    if ok2:
                                        show_snackbar(f"Offline: {info2}")
                                        try:
                                            update_status(False, f"Offline: {info2}")
                                        except Exception:
                                            pass
                                except Exception:
                                    pass
                    except asyncio.TimeoutError:
                        print("[STARTUP] background startup_restore timeout")
                        # Intentar fallback a backup
                        try:
                            ok3, info3 = await asyncio.to_thread(restore_latest_backup)
                            if ok3:
                                show_snackbar(f"Offline: {info3}")
                                try:
                                    update_status(False, f"Offline: {info3}")
                                except Exception:
                                    pass
                        except Exception:
                            pass
                    except Exception as ex2:
                        print(f"[STARTUP] error en background_restore: {ex2}")
                except Exception:
                    pass

            try:
                asyncio.create_task(background_restore())
            except Exception as ex:
                print(f"[STARTUP] no se pudo lanzar background_restore: {ex}")

            # Refrescar lista y estado localmente (sin bloquear)
            try:
                refresh_lotes_list_radios()
            except Exception:
                pass
            try:
                check_and_update_connection_status()
            except Exception:
                pass

            # Mostrar diálogo si falta usuario
            try:
                if not (CURRENT_USER and CURRENT_USER.strip()):
                    mostrar_dialogo_usuario()
            except Exception:
                pass
        asyncio.create_task(startup())

    page.on_load = on_page_load
    
    # ========== DIÁLOGO DE IDENTIFICACIÓN DE USUARIO ==========
    def mostrar_dialogo_usuario():
        """Muestra diálogo para identificar al usuario si no está configurado."""
        if CURRENT_USER:
            return  # Ya hay usuario configurado
        
        nombre_input = ft.TextField(
            label="Tu nombre",
            hint_text="Ej: Eduardo",
            autofocus=True,
            capitalization=ft.TextCapitalization.WORDS,
        )
        
        recordar_check = ft.Checkbox(label="Recordar en este equipo", value=True)
        
        def guardar_y_cerrar(e):
            nombre = nombre_input.value.strip()
            if not nombre:
                nombre_input.error_text = "Ingresa tu nombre"
                page.update()
                return
            
            nombre_normalizado = normalizar_nombre(nombre)
            
            if recordar_check.value:
                guardar_usuario(nombre_normalizado)
            else:
                global CURRENT_USER
                CURRENT_USER = nombre_normalizado
            
            dialogo.open = False
            page.update()
        
        dialogo = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.PERSON, color=ft.Colors.GREEN_700),
                ft.Text("Identificación", weight=ft.FontWeight.BOLD),
            ]),
            content=ft.Column([
                ft.Text("¿Quién está usando la app?", size=14),
                ft.Text("Tu nombre aparecerá en los commits de GitHub.", 
                       size=12, color=ft.Colors.GREY_600),
                nombre_input,
                recordar_check,
            ], tight=True, spacing=10),
            actions=[
                ft.TextButton("Continuar", on_click=guardar_y_cerrar),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        page.overlay.append(dialogo)
        dialogo.open = True
        page.update()
    
    # Mostrar diálogo de usuario al inicio si no hay usuario
    # (se llama después de page.add() para que funcione correctamente)
    
    # Estado - usando controles directos en lugar de Ref
    connection_status = ft.Ref[ft.Container]()
    status_text = ft.Ref[ft.Text]()
    
    # Controles directos para variedades (más confiable que Ref)
    lote_info_label = ft.Text(value="Selecciona un lote", size=12, color=ft.Colors.GREY_700)
    varieties_listview = ft.ListView(
        spacing=0, 
        height=350, 
        padding=0,
        # Mostrar barra de scroll siempre visible
    )
    total_label = ft.Text(value="TOTAL: 0 plantas", size=14, weight=ft.FontWeight.BOLD)
    
    # Variable para guardar el valor seleccionado
    selected_lote = {"value": None}
    
    def show_snackbar(message: str, error: bool = False):
        """Muestra un snackbar con mensaje."""
        page.snack_bar = ft.SnackBar(
            content=ft.Text(message),
            bgcolor=ft.Colors.RED_400 if error else ft.Colors.GREEN_700,
        )
        page.snack_bar.open = True
        page.update()
    
    def update_status(connected: bool, message: str):
        if connection_status.current:
            connection_status.current.bgcolor = ft.Colors.GREEN if connected else ft.Colors.RED
        if status_text.current:
            status_text.current.value = message
        page.update()

    def check_and_update_connection_status():
        """Valida los datos de configuración y actualiza el estado con un mensaje claro."""
        # Priorizar mensajes de error específicos
        if not GITHUB_TOKEN:
            update_status(False, "Sin token configurado")
            return False, "Sin token configurado"
        if not GITHUB_REPO or "/" not in GITHUB_REPO:
            update_status(False, "Repo no configurado")
            return False, "Repo no configurado"
        if not CURRENT_USER:
            update_status(False, "Usuario no configurado")
            return False, "Usuario no configurado"
        # Si todo OK
        update_status(True, "Conectado a GitHub")
        return True, "Conectado a GitHub"
    
    def refresh_lotes_dropdown():
        # Ahora usa los radios en lugar del dropdown
        refresh_lotes_list_radios()
    
    def check_connection(e=None):
        update_status(False, "Verificando...")
        success, msg = descargar_csv_github()
        update_status(success, msg)
        if success:
            refresh_lotes_dropdown()
    
    def sync_to_github(e, manual=True):
        """Sincroniza con GitHub. Si manual=True y hay conflicto, muestra diálogo para resolver."""
        async def do_sync():
            update_status(False, "Sincronizando...")
            # Si el usuario manualmente inició sincronización, comprobar si las subidas están bloqueadas
            try:
                blocked = globals().get('LOCAL_DATA_CLEARED') or os.path.exists(NO_AUTO_RESTORE_FILE)
            except Exception:
                blocked = False

            if manual and blocked:
                # No permitir reactivar subidas desde la app por seguridad
                show_snackbar('Subidas bloqueadas tras borrar datos locales. Eliminar manualmente .no_auto_restore para reactivar si estás seguro.', error=True)
                update_status(False, 'Subidas bloqueadas')
                return
            # Obtener remoto
            ok, msg, remote_content, remote_hash = await asyncio.to_thread(get_remote_csv_content)

            # Si no hay remoto (archivo no encontrado), NO crear automáticamente en auto-sync
            if not ok and isinstance(msg, str) and ('Archivo' in msg or 'Repositorio' in msg):
                if not manual:
                    show_snackbar(f"No se encontró el archivo remoto: {msg}", error=True)
                    update_status(False, msg)
                    return
                else:
                    # En manual, no creamos remotamente automáticamente desde la app
                    show_snackbar('Archivo remoto no encontrado: la app no crea archivos remotos automáticamente. Crea el archivo en GitHub o restaura desde backup manualmente.', error=True)
                    update_status(False, msg)
                    return

            # Leer local
            local_content = ''
            try:
                with open(LOTES_CSV, 'r', encoding='utf-8') as f:
                    local_content = f.read()
            except Exception:
                local_content = ''
            local_hash = compute_hash(local_content) if local_content else ''

            # Si hay diferencia
            if remote_hash and remote_content and remote_content != local_content:
                if not manual:
                    # En auto-sync, no sobrescribimos automáticamente: retornar conflicto
                    show_snackbar('Conflicto remoto: no sincronizado', error=True)
                    update_status(False, 'Conflicto remoto')
                    return

                # Manual: mostrar diálogo con opciones
                def cerrar(e):
                    dlg.open = False
                    page.update()

                def forzar(e):
                    dlg.open = False
                    page.update()
                    async def do_force():
                        success, msg2 = await asyncio.to_thread(lambda: subir_csv_github(force=True))
                        update_status(success, msg2)
                        if not success:
                            show_snackbar(f'Error sincronizando: {msg2}', error=True)
                        else:
                            show_snackbar('Forzado: sincronizado')
                    asyncio.create_task(do_force())

                dlg = ft.AlertDialog(
                    modal=True,
                    title=ft.Text('Conflicto de sincronización'),
                    content=ft.Text('El repositorio remoto ha cambiado. Se ha detenido la sincronización para evitar sobrescribir.'),
                    actions=[
                        ft.TextButton('Cancelar', on_click=cerrar),
                        ft.TextButton('Forzar subir', on_click=forzar, style=ft.ButtonStyle(color=ft.Colors.RED)),
                    ],
                    actions_alignment=ft.MainAxisAlignment.END,
                )
                page.overlay.append(dlg)
                dlg.open = True
                page.update()
                return

            # No hay diferencia/conflicto -> proceder a subir en background
            # Verificar que el CSV local tenga datos antes de subir
            try:
                import io
                with open(LOTES_CSV, 'r', encoding='utf-8') as f:
                    txt = f.read()
                reader = csv.reader(io.StringIO(txt))
                rows = list(reader)
                data_rows = [r for r in rows[1:] if any((c or '').strip() for c in r)] if len(rows) > 1 else []
            except Exception:
                data_rows = None

            if data_rows == []:
                # Mostrar diálogo para confirmar forzar subida de CSV vacío
                def cancelar_force(e):
                    dlg_force.open = False
                    page.update()

                def confirmar_force(e):
                    dlg_force.open = False
                    page.update()
                    async def do_force_upload():
                        success, msg = await asyncio.to_thread(lambda: subir_csv_github(force=True))
                        update_status(success, msg)
                        if not success:
                            show_snackbar(f"Error sincronizando: {msg}", error=True)
                        else:
                            show_snackbar('Sincronizado (forzado)')
                    asyncio.create_task(do_force_upload())

                dlg_force = ft.AlertDialog(
                    modal=True,
                    title=ft.Text('Subir CSV vacío?'),
                    content=ft.Text('El CSV local no contiene datos. Subirlo reemplazará el remoto con un archivo vacío. ¿Deseas continuar?'),
                    actions=[
                        ft.TextButton('Cancelar', on_click=cancelar_force),
                        ft.TextButton('Forzar subir', on_click=confirmar_force, style=ft.ButtonStyle(color=ft.Colors.RED)),
                    ],
                    actions_alignment=ft.MainAxisAlignment.END,
                )
                page.overlay.append(dlg_force)
                dlg_force.open = True
                page.update()
            else:
                success, msg = await asyncio.to_thread(subir_csv_github)
                update_status(success, msg)
                if not success:
                    show_snackbar(f"Error sincronizando: {msg}", error=True)
                else:
                    show_snackbar('Sincronizado exitosamente')

        asyncio.create_task(do_sync())
    
    def create_lote(branch, lote_num, stage, location, semana, notes):
        lotes = leer_csv()
        lotes_rama = [l for l in lotes if l.get('Branch') == branch]
        
        if lote_num == 'AUTO':
            try:
                existing = [int(l.get('LoteNum')) for l in lotes_rama if l.get('LoteNum', '').isdigit()]
                n = max(existing) + 1 if existing else 1
            except:
                n = len(lotes_rama) + 1
        else:
            n = int(lote_num.lstrip('L'))
        
        entry = {
            'ID': f"L{n}-{branch}",
            'Branch': branch,
            'LoteNum': str(n),
            'Stage': stage,
            'Location': location,
            'Semana': str(semana),
            'DateCreated': datetime.now().strftime('%Y-%m-%d'),
            'Notes': notes or '',
            'Variedades': []
        }
        # Evitar crear duplicado exacto (mismo Branch + LoteNum + Location)
        for l in lotes:
            if l.get('Branch') == branch and l.get('LoteNum') == str(n) and l.get('Location') == location:
                # Ya existe un lote con mismo número y ubicación
                return None

        lotes.append(entry)
        guardar_csv(lotes)
        async def do_upload():
            success, msg = await asyncio.to_thread(subir_csv_github)
            if not success:
                show_snackbar(f"⚠️ No sincronizado: {msg}", error=True)
                update_status(False, msg)
            else:
                update_status(True, "Sincronizado")
        asyncio.create_task(do_upload())
        return f"L{n}-{branch}"
    
    def add_variety_to_lote(lote_id, variety_name, qty):
        lotes = leer_csv()
        idx, lote = find_lote_by_id(lote_id, lotes)
        
        if lote is None:
            return False
        
        vars_list = lote.get('Variedades', [])
        found = False
        for v in vars_list:
            if v['name'] == variety_name:
                v['count'] += qty
                found = True
                break
        
        if not found:
            if len(vars_list) >= 20:
                return False
            vars_list.append({'name': variety_name, 'count': qty})
        
        lote['Variedades'] = vars_list
        guardar_csv(lotes)
        async def do_upload():
            success, msg = await asyncio.to_thread(subir_csv_github)
            if not success:
                show_snackbar(f"⚠️ No sincronizado: {msg}", error=True)
                update_status(False, msg)
            else:
                update_status(True, "Sincronizado")
        asyncio.create_task(do_upload())
        return True
    
    def remove_variety_from_lote(lote_id, variety_name):
        lotes = leer_csv()
        idx, lote = find_lote_by_id(lote_id, lotes)
        
        if lote is None:
            return False
        
        vars_list = lote.get('Variedades', [])
        for i, v in enumerate(vars_list):
            if v['name'] == variety_name:
                del vars_list[i]
                lote['Variedades'] = vars_list
                guardar_csv(lotes)
                async def do_upload():
                    success, msg = await asyncio.to_thread(subir_csv_github)
                    if not success:
                        show_snackbar(f"⚠️ No sincronizado: {msg}", error=True)
                        update_status(False, msg)
                    else:
                        update_status(True, "Sincronizado")
                asyncio.create_task(do_upload())
                return True
        return False
    
    # ========== UI COMPONENTS ==========
    
    # Barra de estado
    status_bar = ft.Container(
        content=ft.Row([
            ft.Container(
                ref=connection_status,
                width=12,
                height=12,
                border_radius=6,
                bgcolor=ft.Colors.GREY,
            ),
            ft.Text(ref=status_text, value="Iniciando...", size=12),
            ft.Container(expand=True),
            ft.Text(f"v{VERSION}", size=10, color=ft.Colors.GREY),
        ]),
        padding=10,
        bgcolor=ft.Colors.GREY_100,
        border_radius=8,
    )

    # Botones de sincronización: salen en su propia línea para mejor visualización en móviles
    sync_reconnect_btn = ft.TextButton(content=ft.Text("↻ Reconectar", size=12), on_click=check_connection, style=ft.ButtonStyle(elevation=0))
    sync_upload_btn = ft.TextButton(content=ft.Text("↑ Sincronizar", size=12), on_click=lambda e: sync_to_github(e, manual=True), style=ft.ButtonStyle(elevation=0))

    def build_sync_controls(vertical=False):
        """Construye controles de sincronización en estilo compacto (barra baja)."""
        if vertical:
            return ft.Container(
                content=ft.Column([
                    ft.Container(content=sync_reconnect_btn, height=36, padding=ft.Padding(0, 0)),
                    ft.Container(content=sync_upload_btn, height=36, padding=ft.Padding(0, 0)),
                ], alignment=ft.MainAxisAlignment.END, spacing=4),
                padding=4,
                bgcolor=ft.Colors.GREY_100,
                border_radius=8,
            )
        else:
            return ft.Container(
                content=ft.Row([
                    ft.Container(content=sync_reconnect_btn, height=36, padding=ft.Padding(0, 0)),
                    ft.Container(content=sync_upload_btn, height=36, padding=ft.Padding(0, 0)),
                ], alignment=ft.MainAxisAlignment.END, spacing=4),
                padding=4,
                bgcolor=ft.Colors.GREY_100,
                border_radius=8,
            )

    sync_controls = build_sync_controls()

    def update_sync_layout(e=None):
        try:
            vertical = getattr(page, 'window_width', 800) < 520
            new_ctrl = build_sync_controls(vertical=vertical)
            sync_controls.content = new_ctrl.content
            try:
                sync_controls.update()
            except Exception:
                page.update()
        except Exception:
            pass

    try:
        page.on_resize = update_sync_layout
    except Exception:
        pass

    try:
        update_sync_layout()
    except Exception:
        pass
    
    # ========== TAB 1: CREAR LOTE ==========
    branch_dd = ft.Dropdown(
        label="Sucursal",
        options=[ft.dropdown.Option(b) for b in BRANCH],
        value=None,
        width=220,
    )
    
    lote_num_dd = ft.Dropdown(
        label="Nº Lote",
        options=[ft.dropdown.Option(f"L{i}") for i in range(1, 33)],
        value=None,
        width=140,
    )
    
    stage_dd = ft.Dropdown(
        label="Etapa",
        options=[ft.dropdown.Option(s) for s in STAGES],
        value=None,
        width=240,
    )
    
    location_dd = ft.Dropdown(
        label="Ubicación",
        options=[ft.dropdown.Option(l) for l in LOCATIONS],
        value=None,
        width=220,
    )
    
    semana_dd = ft.Dropdown(
        label="Semana",
        options=[ft.dropdown.Option(str(i)) for i in range(1, 23)],
        value=None,
        width=140,
    )
    
    notes_field = ft.TextField(label="Notas", width=500)
    
    def on_create_click(e):
        # Validar campos requeridos
        if not branch_dd.value or not stage_dd.value or not location_dd.value or not semana_dd.value:
            page.snack_bar = ft.SnackBar(ft.Text("Completa los campos obligatorios: Sucursal, Etapa, Ubicación, Semana"), bgcolor=ft.Colors.RED_400)
            page.snack_bar.open = True
            page.update()
            return

        lote_val = lote_num_dd.value or 'AUTO'
        lote_id = create_lote(
            branch_dd.value,
            lote_val,
            stage_dd.value,
            location_dd.value,
            semana_dd.value,
            notes_field.value
        )
        if not lote_id:
            page.snack_bar = ft.SnackBar(ft.Text("❌ No se creó el lote: ya existe uno igual (mismo número y ubicación)."), bgcolor=ft.Colors.RED_400)
            page.snack_bar.open = True
        else:
            page.snack_bar = ft.SnackBar(ft.Text(f"Lote creado: {lote_id}"))
            page.snack_bar.open = True
            refresh_lotes_dropdown()
            # Limpiar formulario de creación para siguiente ingreso
            try:
                branch_dd.value = None
                lote_num_dd.value = None
                stage_dd.value = None
                location_dd.value = None
                semana_dd.value = None
                notes_field.value = ""
                try:
                    branch_dd.update()
                except Exception:
                    pass
                try:
                    lote_num_dd.update()
                except Exception:
                    pass
                try:
                    stage_dd.update()
                except Exception:
                    pass
                try:
                    location_dd.update()
                except Exception:
                    pass
                try:
                    semana_dd.update()
                except Exception:
                    pass
            except Exception:
                pass
        page.update()
    
    tab_crear = ft.Column([
        ft.Text("Crear nuevo lote", size=20, weight=ft.FontWeight.BOLD),
        ft.Divider(),
        ft.Row([branch_dd, lote_num_dd], wrap=True),
        ft.Row([stage_dd, location_dd], wrap=True),
        ft.Row([semana_dd], wrap=True),
        notes_field,
        ft.FilledButton(
            "Crear Lote",
            icon=ft.Icons.ADD,
            on_click=on_create_click,
            style=ft.ButtonStyle(bgcolor=ft.Colors.GREEN, color=ft.Colors.WHITE),
        ),
    ], spacing=15, scroll=ft.ScrollMode.AUTO)
    
    # ========== TAB 2: AGREGAR VARIEDADES ==========
    variety_dd = ft.Dropdown(
        label="Variedad",
        options=[ft.dropdown.Option(v) for v in sorted(VARIETIES)],
        value=sorted(VARIETIES)[0],
        width=200,
    )
    
    qty_field = ft.TextField(label="Cantidad", value="1", width=100, keyboard_type=ft.KeyboardType.NUMBER)
    
    # Variable para tracking (se define antes de usarse)
    current_lote_id = {"value": None}
    
    def on_add_variety(e):
        if not current_lote_id["value"]:
            page.snack_bar = ft.SnackBar(ft.Text("Selecciona un lote primero"))
            page.snack_bar.open = True
            page.update()
            return
        
        try:
            qty = int(qty_field.value)
        except:
            page.snack_bar = ft.SnackBar(ft.Text("Cantidad inválida"))
            page.snack_bar.open = True
            page.update()
            return
        
        if add_variety_to_lote(current_lote_id["value"], variety_dd.value, qty):
            page.snack_bar = ft.SnackBar(ft.Text(f"Agregado: {variety_dd.value} x{qty}"))
            load_lote_data(current_lote_id["value"])
        else:
            page.snack_bar = ft.SnackBar(ft.Text("Error al agregar"))
        page.snack_bar.open = True
        page.update()
    
    # Usar PopupMenuButton que SÍ tiene on_click funcional
    lote_selector_text = ft.Text("Seleccionar lote...", size=14)
    lotes_popup_menu = ft.PopupMenuButton(
        content=ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.FOLDER_OPEN, size=20),
                lote_selector_text,
                ft.Icon(ft.Icons.ARROW_DROP_DOWN),
            ], spacing=8),
            padding=ft.Padding(left=12, right=12, top=8, bottom=8),
            border=ft.Border.all(1, ft.Colors.GREY_400),
            border_radius=8,
        ),
        items=[],
    )
    
    def on_lote_selected(lote_id):
        """Cuando se selecciona un lote del popup menu."""
        current_lote_id["value"] = lote_id
        lote_selector_text.value = lote_id
        load_lote_data(lote_id)
        page.update()
    
    def refresh_lotes_list_radios():
        """Actualiza la lista de lotes en el popup menu."""
        ids = get_lote_ids_sorted()
        lotes_popup_menu.items.clear()
        for lote_id in ids:
            item = ft.PopupMenuItem(
                content=ft.Text(lote_id),
                on_click=lambda e, lid=lote_id: on_lote_selected(lid),
            )
            lotes_popup_menu.items.append(item)
        if ids:
            current_lote_id["value"] = ids[0]
            lote_selector_text.value = ids[0]
            load_lote_data(ids[0])
        else:
            # No hay lotes locales: limpiar selección y UI
            current_lote_id["value"] = None
            lote_selector_text.value = "No hay lotes locales"
            lote_info_label.value = "No hay lotes locales"
            try:
                varieties_listview.controls.clear()
                varieties_listview.controls.append(ft.Text("Sin datos", color=ft.Colors.GREY_500, italic=True))
            except Exception:
                pass
            total_label.value = "TOTAL: 0 plantas"
        page.update()
    
    def eliminar_variedad(variedad_name):
        """Elimina una variedad del lote actual."""
        if not current_lote_id["value"]:
            return
        
        lotes = leer_csv()
        idx, lote = find_lote_by_id(current_lote_id["value"], lotes)
        
        if lote is None:
            return
        
        vars_list = lote.get('Variedades', [])
        for i, v in enumerate(vars_list):
            if v['name'] == variedad_name:
                del vars_list[i]
                lote['Variedades'] = vars_list
                guardar_csv(lotes)
                async def do_upload():
                    success, msg = await asyncio.to_thread(subir_csv_github)
                    if not success:
                        show_snackbar(f"⚠️ No sincronizado: {msg}", error=True)
                    else:
                        page.snack_bar = ft.SnackBar(ft.Text(f"Eliminado: {variedad_name}"))
                        page.snack_bar.open = True
                    load_lote_data(current_lote_id["value"])
                asyncio.create_task(do_upload())
                return
    
    def confirmar_eliminar(variedad_name):
        """Muestra diálogo de confirmación antes de eliminar."""
        def cerrar_dialogo(e):
            dialogo.open = False
            page.update()
        
        def confirmar(e):
            dialogo.open = False
            page.update()
            eliminar_variedad(variedad_name)
        
        dialogo = ft.AlertDialog(
            modal=True,
            title=ft.Text("¿Eliminar variedad?"),
            content=ft.Text(f"¿Seguro que deseas eliminar '{variedad_name}' del lote?"),
            actions=[
                ft.TextButton("Cancelar", on_click=cerrar_dialogo),
                ft.TextButton("Eliminar", on_click=confirmar, style=ft.ButtonStyle(color=ft.Colors.RED)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        page.overlay.append(dialogo)
        dialogo.open = True
        page.update()
    
    def build_variety_tile(v):
        """Construye un tile de variedad con botón eliminar."""
        return ft.Container(
            content=ft.Row([
                ft.Text("🌿", size=11),
                ft.Text(v['name'], size=12, expand=True),
                ft.Text(str(v['count']), size=12, weight=ft.FontWeight.BOLD, width=35, text_align=ft.TextAlign.RIGHT),
                ft.GestureDetector(
                    content=ft.Container(
                        content=ft.Text("✕", size=16, color=ft.Colors.RED_400, weight=ft.FontWeight.BOLD),
                        padding=ft.Padding(left=8, right=4, top=0, bottom=0),
                    ),
                    data=v['name'],
                    on_tap=lambda e: confirmar_eliminar(e.control.data),
                ),
            ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            padding=ft.Padding(left=8, right=4, top=4, bottom=4),
            border=ft.Border(bottom=ft.BorderSide(1, ft.Colors.GREY_200)),
        )
    
    def load_lote_data(sel):
        """Carga los datos de un lote específico."""
        if not sel:
            return
        
        current_lote_id["value"] = sel
        
        # Siempre leer datos frescos del CSV
        lotes_data = leer_csv()
        idx, lote = find_lote_by_id(sel, lotes_data)
        
        if lote is None:
            lote_info_label.value = "Lote no encontrado"
            varieties_listview.controls.clear()
            total_label.value = "TOTAL: 0 plantas"
            page.update()
            return
        
        # Actualizar info
        lote_info_label.value = f"🌱 {lote.get('Stage', '')} | 📍 {lote.get('Location', '')} | 📅 Sem. {lote.get('Semana', '')}"
        
        # Actualizar lista de variedades
        variedades = lote.get('Variedades', [])
        total = sum(v['count'] for v in variedades)
        
        varieties_listview.controls.clear()
        if not variedades:
            varieties_listview.controls.append(
                ft.Text("Sin variedades", color=ft.Colors.GREY_500, italic=True)
            )
        else:
            for v in sorted(variedades, key=lambda x: x['name']):
                varieties_listview.controls.append(build_variety_tile(v))
        
        total_label.value = f"🌿 TOTAL: {total} plantas"
        page.update()

    tab_variedades = ft.Column([
        ft.Text("Agregar variedades", size=20, weight=ft.FontWeight.BOLD),
        ft.Divider(),
        ft.Text("Selecciona un lote:", size=12),
        ft.Row([
            lotes_popup_menu,
            ft.IconButton(ft.Icons.REFRESH, on_click=lambda e: refresh_lotes_list_radios(), tooltip="Refrescar lista"),
        ]),
        lote_info_label,
        ft.Container(
            content=varieties_listview,
            border=ft.Border.all(1, ft.Colors.GREY_300),
            border_radius=8,
            padding=5,
        ),
        total_label,
        ft.Divider(),
        ft.Row([variety_dd, qty_field], wrap=True),
        ft.FilledButton(
            "Agregar variedad",
            icon=ft.Icons.ADD_CIRCLE,
            on_click=on_add_variety,
            style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE, color=ft.Colors.WHITE),
        ),
    ], spacing=10, scroll=ft.ScrollMode.AUTO)
    
    # ========== TAB 3: GRÁFICOS ==========
    
    def build_stage_chart():
        """Construye visualización de distribución por etapa usando barras."""
        lotes = leer_csv()
        por_etapa = {}
        
        for lote in lotes:
            stage = lote.get('Stage', 'Sin etapa')
            por_etapa[stage] = por_etapa.get(stage, 0) + 1
        
        if not por_etapa:
            return ft.Text("No hay datos para mostrar")
        
        total = sum(por_etapa.values())
        colors = [
            ft.Colors.GREEN_400, ft.Colors.BLUE_400, ft.Colors.PURPLE_400,
            ft.Colors.ORANGE_400, ft.Colors.RED_400, ft.Colors.TEAL_400, ft.Colors.PINK_400
        ]
        
        # Usar barras de progreso visuales en lugar de PieChart
        items = []
        max_count = max(por_etapa.values())
        
        for i, (stage, count) in enumerate(sorted(por_etapa.items(), key=lambda x: -x[1])):
            pct = (count / total) * 100
            bar_width = (count / max_count)
            
            items.append(
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Text(stage, size=12, weight=ft.FontWeight.BOLD, width=120),
                            ft.Container(
                                width=150 * bar_width,
                                height=20,
                                bgcolor=colors[i % len(colors)],
                                border_radius=4,
                            ),
                            ft.Text(f" {count} ({pct:.0f}%)", size=12),
                        ], spacing=10),
                    ]),
                    padding=5,
                )
            )
        
        items.append(ft.Divider())
        items.append(ft.Text(f"Total: {total} lotes", weight=ft.FontWeight.BOLD))
        
        return ft.Column(items, spacing=5)
    
    def build_location_chart():
        """Construye visualización por ubicación."""
        lotes = leer_csv()
        por_ubicacion = {}
        
        for lote in lotes:
            loc = lote.get('Location', 'Sin ubicación')
            por_ubicacion[loc] = por_ubicacion.get(loc, 0) + 1
        
        if not por_ubicacion:
            return ft.Text("No hay datos para mostrar")
        
        max_val = max(por_ubicacion.values())
        total = sum(por_ubicacion.values())
        
        items = []
        for loc, count in sorted(por_ubicacion.items(), key=lambda x: -x[1]):
            bar_width = (count / max_val)
            items.append(
                ft.Row([
                    ft.Text(loc, size=11, width=100),
                    ft.Container(
                        width=180 * bar_width,
                        height=18,
                        bgcolor=ft.Colors.BLUE_400,
                        border_radius=4,
                    ),
                    ft.Text(f" {count}", size=11, weight=ft.FontWeight.BOLD),
                ], spacing=8)
            )
        
        items.append(ft.Divider())
        items.append(ft.Text(f"Total: {total} lotes", weight=ft.FontWeight.BOLD))
        
        return ft.Column(items, spacing=8)
    
    def build_branch_chart():
        """Construye visualización por sucursal y etapa."""
        lotes = leer_csv()
        data = {}
        
        for lote in lotes:
            branch = lote.get('Branch', 'N/A')
            stage = lote.get('Stage', 'N/A')
            key = (branch, stage)
            data[key] = data.get(key, 0) + 1
        
        if not data:
            return ft.Text("No hay datos para mostrar")
        
        branches = sorted(list(set(k[0] for k in data.keys())))
        stages_order = ['CLONADO', 'VEG. TEMPRANO', 'VEG. TARDIO', 'FLORACIÓN', 'TRANSICIÓN', 'SECADO', 'PT']
        
        stage_colors = {
            'CLONADO': ft.Colors.GREEN_300,
            'VEG. TEMPRANO': ft.Colors.GREEN_500,
            'VEG. TARDIO': ft.Colors.GREEN_700,
            'FLORACIÓN': ft.Colors.PURPLE_400,
            'TRANSICIÓN': ft.Colors.ORANGE_400,
            'SECADO': ft.Colors.BROWN_400,
            'PT': ft.Colors.GREY_400,
        }
        
        items = []
        
        for branch in branches:
            branch_row = [ft.Text(branch, size=14, weight=ft.FontWeight.BOLD, width=50)]
            
            for stage in stages_order:
                count = data.get((branch, stage), 0)
                if count > 0:
                    branch_row.append(
                        ft.Container(
                            content=ft.Text(str(count), size=10, color=ft.Colors.WHITE),
                            bgcolor=stage_colors.get(stage, ft.Colors.BLUE_400),
                            padding=ft.padding.symmetric(horizontal=8, vertical=4),
                            border_radius=4,
                            tooltip=f"{stage}: {count}",
                        )
                    )
            
            items.append(ft.Row(branch_row, spacing=5, wrap=True))
        
        # Leyenda
        legend = ft.Row([
            ft.Container(
                content=ft.Row([
                    ft.Container(width=12, height=12, bgcolor=color, border_radius=2),
                    ft.Text(stage[:10], size=9),
                ], spacing=2),
                padding=2,
            )
            for stage, color in stage_colors.items()
        ], wrap=True, spacing=3)
        
        items.append(ft.Divider())
        items.append(legend)
        
        return ft.Column(items, spacing=10)
    
    chart_container = ft.Ref[ft.Container]()
    
    def show_chart(chart_type):
        if chart_container.current:
            if chart_type == "etapas":
                chart_container.current.content = build_stage_chart()
            elif chart_type == "ubicaciones":
                chart_container.current.content = build_location_chart()
            elif chart_type == "sucursales":
                chart_container.current.content = build_branch_chart()
            page.update()
    
    tab_graficos = ft.Column([
        ft.Text("Gráficos", size=20, weight=ft.FontWeight.BOLD),
        ft.Divider(),
        ft.Row([
            ft.FilledButton("Por Etapa", icon=ft.Icons.PIE_CHART, 
                            on_click=lambda e: show_chart("etapas")),
            ft.FilledButton("Por Ubicación", icon=ft.Icons.BAR_CHART,
                            on_click=lambda e: show_chart("ubicaciones")),
            ft.FilledButton("Por Sucursal", icon=ft.Icons.STACKED_BAR_CHART,
                            on_click=lambda e: show_chart("sucursales")),
        ], wrap=True),
        ft.Container(
            ref=chart_container,
            content=ft.Text("Selecciona un gráfico", color=ft.Colors.GREY),
            height=350,
            border=ft.Border.all(1, ft.Colors.GREY_300),
            border_radius=8,
            padding=10,
        ),
    ], spacing=15, scroll=ft.ScrollMode.AUTO)
    
    # ========== TAB 4: LISTADO ==========
    lotes_listview = ft.Ref[ft.ListView]()
    
    # Funciones de exportación
    def get_export_data():
        """Obtiene los datos filtrados para exportar"""
        lotes = leer_csv()
        
        # Aplicar filtros actuales
        branch_filter = filter_branch_dd.value if filter_branch_dd.value != "Todas" else None
        stage_filter = filter_stage_dd.value if filter_stage_dd.value != "Todas" else None
        location_filter = filter_location_dd.value if filter_location_dd.value != "Todas" else None
        
        filtered = []
        for lote in lotes:
            if branch_filter and lote.get('Branch') != branch_filter:
                continue
            if stage_filter and lote.get('Stage') != stage_filter:
                continue
            if location_filter and lote.get('Location') != location_filter:
                continue
            filtered.append(lote)
        
        def lote_key(lote):
            try:
                return (lote.get('Branch', ''), int(lote.get('LoteNum', 0)))
            except:
                return (lote.get('Branch', ''), 0)
        
        return sorted(filtered, key=lote_key)
    
    def get_downloads_folder():
        """Obtiene la carpeta de Descargas según el sistema operativo."""
        if sys.platform == 'android':
            return "/storage/emulated/0/Download"
        # Linux/Mac: ~/Descargas o ~/Downloads
        home = os.path.expanduser("~")
        for folder in ["Descargas", "Downloads"]:
            path = os.path.join(home, folder)
            if os.path.isdir(path):
                return path
        # Si no existe, crear Downloads
        path = os.path.join(home, "Downloads")
        os.makedirs(path, exist_ok=True)
        return path
    
    def show_export_success(filepath, file_type):
        """Muestra diálogo de éxito con la ruta del archivo exportado."""
        def cerrar_dialogo(e):
            dlg.open = False
            page.update()
        
        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text(f"✅ {file_type} exportado"),
            content=ft.Column([
                ft.Text("Archivo guardado en:", size=12),
                ft.Container(
                    content=ft.Text(filepath, size=11, selectable=True),
                    bgcolor=ft.Colors.GREY_200,
                    padding=10,
                    border_radius=5,
                ),
            ], tight=True, spacing=10),
            actions=[ft.TextButton("OK", on_click=cerrar_dialogo)],
        )
        page.overlay.append(dlg)
        dlg.open = True
        page.update()
    
    def export_to_csv(e=None):
        """Exportar a CSV"""
        lotes = get_export_data()
        if not lotes:
            show_snackbar("No hay datos para exportar", error=True)
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"lotes_export_{timestamp}.csv"
        
        # Usar carpeta de Descargas
        export_dir = get_downloads_folder()
        filepath = os.path.join(export_dir, filename)
        
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                # Encabezados
                writer.writerow(['ID', 'Sucursal', 'Lote', 'Etapa', 'Ubicación', 'Semana', 'Fecha', 'Variedades', 'Total Plantas', 'Notas'])
                
                for lote in lotes:
                    variedades = lote.get('Variedades', [])
                    vars_str = ', '.join([f"{v['name']}({v['count']})" for v in variedades])
                    total = sum(v['count'] for v in variedades)
                    
                    writer.writerow([
                        lote.get('ID', ''),
                        lote.get('Branch', ''),
                        lote.get('LoteNum', ''),
                        lote.get('Stage', ''),
                        lote.get('Location', ''),
                        lote.get('Semana', ''),
                        lote.get('DateCreated', ''),
                        vars_str,
                        total,
                        lote.get('Notes', '')
                    ])
            
            show_export_success(filepath, "CSV")
        except Exception as ex:
            show_snackbar(f"Error al exportar CSV: {ex}", error=True)
    
    def export_to_excel(e=None):
        """Exportar a Excel (XLSX)"""
        if not OPENPYXL_AVAILABLE:
            show_snackbar("⚠️ openpyxl no disponible. Instalar: pip install openpyxl", error=True)
            return
        
        lotes = get_export_data()
        if not lotes:
            show_snackbar("No hay datos para exportar", error=True)
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"lotes_export_{timestamp}.xlsx"
        
        # Usar carpeta de Descargas
        export_dir = get_downloads_folder()
        filepath = os.path.join(export_dir, filename)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Lotes"
            
            # Encabezados con formato
            headers = ['ID', 'Sucursal', 'Lote', 'Etapa', 'Ubicación', 'Semana', 'Fecha', 'Variedades', 'Total', 'Notas']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Datos
            for row_num, lote in enumerate(lotes, 2):
                variedades = lote.get('Variedades', [])
                vars_str = ', '.join([f"{v['name']}({v['count']})" for v in variedades])
                total = sum(v['count'] for v in variedades)
                
                ws.cell(row=row_num, column=1, value=lote.get('ID', ''))
                ws.cell(row=row_num, column=2, value=lote.get('Branch', ''))
                ws.cell(row=row_num, column=3, value=int(lote.get('LoteNum', 0)) if lote.get('LoteNum', '').isdigit() else 0)
                ws.cell(row=row_num, column=4, value=lote.get('Stage', ''))
                ws.cell(row=row_num, column=5, value=lote.get('Location', ''))
                ws.cell(row=row_num, column=6, value=int(lote.get('Semana', 0)) if lote.get('Semana', '').isdigit() else 0)
                ws.cell(row=row_num, column=7, value=lote.get('DateCreated', ''))
                ws.cell(row=row_num, column=8, value=vars_str)
                ws.cell(row=row_num, column=9, value=total)
                ws.cell(row=row_num, column=10, value=lote.get('Notes', ''))
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['J'].width = 25
            
            wb.save(filepath)
            show_export_success(filepath, "Excel")
        except Exception as ex:
            show_snackbar(f"Error al exportar Excel: {ex}", error=True)
    
    def export_to_pdf(e=None):
        """Exportar a PDF con todas las variedades visibles"""
        if not FPDF_AVAILABLE:
            show_snackbar("⚠️ fpdf2 no disponible. Instalar: pip install fpdf2", error=True)
            return
        
        lotes = get_export_data()
        if not lotes:
            show_snackbar("No hay datos para exportar", error=True)
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"lotes_export_{timestamp}.pdf"
        
        # Usar carpeta de Descargas
        export_dir = get_downloads_folder()
        filepath = os.path.join(export_dir, filename)
        
        try:
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            
            # Título
            pdf.set_font('Helvetica', 'B', 18)
            pdf.cell(0, 12, 'Control de Lotes - Reporte', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 8, f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            pdf.ln(5)
            
            # Filtros aplicados
            filters = []
            if filter_branch_dd.value != "Todas":
                filters.append(f"Sucursal: {filter_branch_dd.value}")
            if filter_stage_dd.value != "Todas":
                filters.append(f"Etapa: {filter_stage_dd.value}")
            if filter_location_dd.value != "Todas":
                filters.append(f"Ubicación: {filter_location_dd.value}")
            if filters:
                pdf.set_font('Helvetica', 'I', 9)
                pdf.cell(0, 6, f'Filtros: {" | ".join(filters)}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            pdf.ln(5)
            
            # Formato de ficha por cada lote
            for lote in lotes:
                variedades = lote.get('Variedades', [])
                total = sum(v['count'] for v in variedades)
                
                # Verificar si hay espacio suficiente, sino nueva página
                needed_height = 30 + (len(variedades) * 5)
                if pdf.get_y() + needed_height > 270:
                    pdf.add_page()
                
                # Encabezado del lote (fondo gris)
                pdf.set_fill_color(230, 230, 230)
                pdf.set_font('Helvetica', 'B', 11)
                lote_id = lote.get('ID', '')
                stage = lote.get('Stage', '')
                location = lote.get('Location', '')
                semana = lote.get('Semana', '')
                
                pdf.cell(0, 8, f"{lote_id}  |  {stage}  |  {location}  |  Semana {semana}  |  Total: {total} plantas", 
                         new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True, border=1)
                
                # Lista de variedades
                if variedades:
                    pdf.set_font('Helvetica', '', 9)
                    for v in sorted(variedades, key=lambda x: x['name']):
                        pdf.cell(10, 5, '', border=0)  # Indentación
                        pdf.cell(80, 5, f"- {v['name']}", border=0)
                        pdf.cell(30, 5, str(v['count']), new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
                else:
                    pdf.set_font('Helvetica', 'I', 9)
                    pdf.cell(10, 5, '', border=0)
                    pdf.cell(0, 5, 'Sin variedades', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                pdf.ln(3)
            
            # Resumen final
            pdf.ln(5)
            pdf.set_draw_color(0, 0, 0)
            pdf.set_font('Helvetica', 'B', 12)
            total_lotes = len(lotes)
            total_plantas = sum(sum(v['count'] for v in l.get('Variedades', [])) for l in lotes)
            pdf.cell(0, 10, f'TOTAL: {total_lotes} lotes  |  {total_plantas} plantas', 
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=1, align='C')
            
            pdf.output(filepath)
            show_export_success(filepath, "PDF")
        except ImportError as ie:
            show_snackbar(f"Error de importación: {ie}", error=True)
        except Exception as ex:
            import traceback
            print(f"Error PDF: {traceback.format_exc()}")
            show_snackbar(f"Error al exportar PDF: {ex}", error=True)
    
    # Filtros para el listado
    filter_branch_dd = ft.Dropdown(
        label="Sucursal",
        options=[ft.dropdown.Option("Todas")] + [ft.dropdown.Option(b) for b in BRANCH],
        value="Todas",
        width=120,
        dense=True,
    )
    filter_stage_dd = ft.Dropdown(
        label="Etapa",
        options=[ft.dropdown.Option("Todas")] + [ft.dropdown.Option(s) for s in STAGES],
        value="Todas",
        width=150,
        dense=True,
    )
    filter_location_dd = ft.Dropdown(
        label="Ubicación",
        options=[ft.dropdown.Option("Todas")] + [ft.dropdown.Option(l) for l in LOCATIONS],
        value="Todas",
        width=140,
        dense=True,
    )
    
    def refresh_lotes_list(e=None):
        lotes = leer_csv()
        
        # Aplicar filtros
        branch_filter = filter_branch_dd.value if filter_branch_dd.value != "Todas" else None
        stage_filter = filter_stage_dd.value if filter_stage_dd.value != "Todas" else None
        location_filter = filter_location_dd.value if filter_location_dd.value != "Todas" else None
        
        filtered = []
        for lote in lotes:
            if branch_filter and lote.get('Branch') != branch_filter:
                continue
            if stage_filter and lote.get('Stage') != stage_filter:
                continue
            if location_filter and lote.get('Location') != location_filter:
                continue
            filtered.append(lote)
        
        def lote_key(lote):
            try:
                return (lote.get('Branch', ''), int(lote.get('LoteNum', 0)))
            except:
                return (lote.get('Branch', ''), 0)
        
        lotes_sorted = sorted(filtered, key=lote_key)
        
        if lotes_listview.current:
            lotes_listview.current.controls.clear()
            if not lotes_sorted:
                # Mostrar mensaje claro cuando no hay datos
                lotes_listview.current.controls.append(ft.Text("No hay lotes locales", color=ft.Colors.GREY_600))
            else:
                for lote in lotes_sorted:
                    branch = lote.get('Branch', '')
                    lote_num = lote.get('LoteNum', '')
                    lote_id = f"L{lote_num}-{branch}"
                    
                    variedades = lote.get('Variedades', [])
                    total = sum(v['count'] for v in variedades)
                    
                    # Mostrar todas las variedades en líneas separadas
                    vars_widgets = []
                    if variedades:
                        for v in sorted(variedades, key=lambda x: x['name']):
                            vars_widgets.append(
                                ft.Text(f"  🌿 {v['name']}: {v['count']}", size=11, color=ft.Colors.GREY_700)
                            )
                    else:
                        vars_widgets.append(ft.Text("  Sin variedades", size=11, color=ft.Colors.GREY_500, italic=True))
                    
                    lotes_listview.current.controls.append(
                        ft.Card(
                            content=ft.Container(
                                content=ft.Column([
                                    ft.Row([
                                        ft.Text(lote_id, size=16, weight=ft.FontWeight.BOLD),
                                        ft.Container(expand=True),
                                        ft.Chip(label=ft.Text(lote.get('Stage', '')), bgcolor=ft.Colors.GREEN_100),
                                    ]),
                                    ft.Text(f"📍 {lote.get('Location', '')} | 📅 Semana {lote.get('Semana', '')}", size=12),
                                    ft.Column(vars_widgets, spacing=0),
                                    ft.Text(f"🌱 Total: {total} plantas", size=12, weight=ft.FontWeight.W_500),
                                ], spacing=4),
                                padding=12,
                            ),
                        )
                    )
            page.update()
    
    tab_listado = ft.Column([
        ft.Row([
            ft.Text("Listado de Lotes", size=20, weight=ft.FontWeight.BOLD),
            ft.Container(expand=True),
            ft.IconButton(ft.Icons.REFRESH, on_click=refresh_lotes_list, tooltip="Actualizar"),
        ]),
        ft.Row([
            filter_branch_dd,
            filter_stage_dd,
            filter_location_dd,
            ft.FilledButton("Filtrar", icon=ft.Icons.FILTER_ALT, on_click=refresh_lotes_list),
            ft.TextButton("Limpiar", on_click=lambda e: clear_filters()),
        ], wrap=True, spacing=8),
        ft.Divider(),
        # Botones de exportar
        ft.Row([
            ft.Text("Exportar:", size=12, color=ft.Colors.GREY_600),
            ft.OutlinedButton("CSV", icon=ft.Icons.TABLE_CHART, on_click=export_to_csv),
            ft.OutlinedButton("Excel", icon=ft.Icons.GRID_ON, on_click=export_to_excel),
            ft.OutlinedButton("PDF", icon=ft.Icons.PICTURE_AS_PDF, on_click=export_to_pdf),
        ], spacing=8),
        ft.Divider(),
        ft.ListView(ref=lotes_listview, spacing=8, expand=True),
    ], expand=True)
    
    def clear_filters():
        filter_branch_dd.value = "Todas"
        filter_stage_dd.value = "Todas"
        filter_location_dd.value = "Todas"
        refresh_lotes_list()
        page.update()
    
    # ========== TAB 5: EDITAR LOTE ==========
    edit_lote_selector_text = ft.Text("Seleccionar lote...", size=14)
    edit_lote_popup = ft.PopupMenuButton(
        content=ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.EDIT, size=20),
                edit_lote_selector_text,
                ft.Icon(ft.Icons.ARROW_DROP_DOWN),
            ], spacing=8),
            padding=ft.Padding(left=12, right=12, top=8, bottom=8),
            border=ft.Border.all(1, ft.Colors.GREY_400),
            border_radius=8,
        ),
        items=[],
    )
    
    current_edit_lote = {"value": None}
    edit_info_label = ft.Text("", size=12, color=ft.Colors.GREY_600)
    # Ref para el botón de guardar en edición para poder habilitar/deshabilitar
    edit_save_btn_ref = ft.Ref[ft.FilledButton]()
    
    edit_stage_dd = ft.Dropdown(
        label="Nueva Etapa",
        options=[ft.dropdown.Option(s) for s in STAGES],
        hint_text="",
        width=200,
    )
    
    edit_location_dd = ft.Dropdown(
        label="Nueva Ubicación",
        options=[ft.dropdown.Option(l) for l in LOCATIONS],
        hint_text="",
        width=200,
    )
    
    edit_semana_dd = ft.Dropdown(
        label="Nueva Semana",
        options=[ft.dropdown.Option(str(i)) for i in range(1, 23)],
        hint_text="",
        width=120,
    )
    
    def on_edit_lote_selected(lote_id):
        """Cuando se selecciona un lote para editar."""
        current_edit_lote["value"] = lote_id
        edit_lote_selector_text.value = lote_id
        
        # Cargar datos del lote
        lotes = leer_csv()
        idx, lote = find_lote_by_id(lote_id, lotes)
        
        if lote:
            edit_stage_dd.value = lote.get('Stage', '')
            edit_location_dd.value = lote.get('Location', '')
            edit_semana_dd.value = lote.get('Semana', '')
            edit_info_label.value = f"Editando: {lote_id}"
            # Habilitar dropdowns y boton de guardar
            try:
                edit_stage_dd.disabled = False
                edit_location_dd.disabled = False
                edit_semana_dd.disabled = False
                try:
                    edit_stage_dd.update()
                except Exception:
                    pass
                try:
                    edit_location_dd.update()
                except Exception:
                    pass
                try:
                    edit_semana_dd.update()
                except Exception:
                    pass
            except Exception:
                pass
            try:
                if edit_save_btn_ref and edit_save_btn_ref.current:
                    edit_save_btn_ref.current.disabled = False
                    edit_save_btn_ref.current.update()
            except Exception:
                pass
        else:
            edit_info_label.value = "Lote no encontrado"
            try:
                # Dejar dropdowns vacíos pero habilitados para nuevas selecciones
                edit_stage_dd.value = ""
                edit_location_dd.value = ""
                edit_semana_dd.value = ""
                edit_stage_dd.disabled = False
                edit_location_dd.disabled = False
                edit_semana_dd.disabled = False
                try:
                    edit_stage_dd.update()
                except Exception:
                    pass
                try:
                    edit_location_dd.update()
                except Exception:
                    pass
                try:
                    edit_semana_dd.update()
                except Exception:
                    pass
            except Exception:
                pass
            try:
                if edit_save_btn_ref and edit_save_btn_ref.current:
                    edit_save_btn_ref.current.disabled = True
                    edit_save_btn_ref.current.update()
            except Exception:
                pass
        
        page.update()
    
    def refresh_edit_lotes_popup():
        """Actualiza la lista de lotes en el popup de edición."""
        ids = get_lote_ids_sorted()
        edit_lote_popup.items.clear()
        for lote_id in ids:
            item = ft.PopupMenuItem(
                content=ft.Text(lote_id),
                on_click=lambda e, lid=lote_id: on_edit_lote_selected(lid),
            )
            edit_lote_popup.items.append(item)
        if ids:
            # Si no hay selección actual, seleccionar la primera
            if not current_edit_lote["value"]:
                on_edit_lote_selected(ids[0])
        else:
            # No hay lotes: limpiar controles de edición
            current_edit_lote["value"] = None
            edit_lote_selector_text.value = "No hay lotes"
            edit_info_label.value = "No hay lotes para editar"
            try:
                edit_stage_dd.value = ""
                edit_location_dd.value = ""
                edit_semana_dd.value = ""
                # Dropdowns vacíos pero habilitados para nuevas selecciones
                edit_stage_dd.disabled = False
                edit_location_dd.disabled = False
                edit_semana_dd.disabled = False
                try:
                    edit_stage_dd.update()
                except Exception:
                    pass
                try:
                    edit_location_dd.update()
                except Exception:
                    pass
                try:
                    edit_semana_dd.update()
                except Exception:
                    pass
            except Exception:
                pass
            try:
                if edit_save_btn_ref and edit_save_btn_ref.current:
                    edit_save_btn_ref.current.disabled = True
                    edit_save_btn_ref.current.update()
            except Exception:
                pass
        page.update()
    
    def on_guardar_edicion(e):
        """Guarda los cambios del lote editado."""
        if not current_edit_lote["value"]:
            page.snack_bar = ft.SnackBar(ft.Text("Selecciona un lote primero"))
            page.snack_bar.open = True
            page.update()
            return
        
        lotes = leer_csv()
        idx, lote = find_lote_by_id(current_edit_lote["value"], lotes)
        
        if lote is None:
            page.snack_bar = ft.SnackBar(ft.Text("Lote no encontrado"))
            page.snack_bar.open = True
            page.update()
            return
        
        cambios = []
        
        # Aplicar cambios
        if edit_stage_dd.value and lote.get('Stage') != edit_stage_dd.value:
            lote['Stage'] = edit_stage_dd.value
            cambios.append('Etapa')
        
        if edit_location_dd.value and lote.get('Location') != edit_location_dd.value:
            lote['Location'] = edit_location_dd.value
            cambios.append('Ubicación')
        
        if edit_semana_dd.value and lote.get('Semana') != edit_semana_dd.value:
            lote['Semana'] = edit_semana_dd.value
            cambios.append('Semana')
            # Si la semana es 20 o 21, cambiar automáticamente a PT/SECADO
            try:
                sem_num = int(edit_semana_dd.value)
                if sem_num in (20, 21):
                    lote['Location'] = 'PT'
                    lote['Stage'] = 'SECADO'
                    if 'Ubicación' not in cambios:
                        cambios.append('Ubicación→PT')
                    if 'Etapa' not in cambios:
                        cambios.append('Etapa→SECADO')
            except:
                pass
        
        if not cambios:
            page.snack_bar = ft.SnackBar(ft.Text("No hay cambios para guardar"))
            page.snack_bar.open = True
            page.update()
            return
        
        # Guardar y sincronizar
        guardar_csv(lotes)
        async def do_upload():
            success, msg = await asyncio.to_thread(subir_csv_github)
            if not success:
                show_snackbar(f"⚠️ No sincronizado: {msg}", error=True)
            else:
                page.snack_bar = ft.SnackBar(ft.Text(f"✅ Lote actualizado ({', '.join(cambios)})"))
                page.snack_bar.open = True
            # Refrescar listas
            refresh_edit_lotes_popup()
            refresh_lotes_dropdown()
            page.update()
        asyncio.create_task(do_upload())
    
    tab_editar = ft.Column([
        ft.Text("Editar Lote", size=20, weight=ft.FontWeight.BOLD),
        ft.Divider(),
        ft.Text("Selecciona un lote:", size=12),
        ft.Row([
            edit_lote_popup,
            ft.IconButton(ft.Icons.REFRESH, on_click=lambda e: refresh_edit_lotes_popup(), tooltip="Refrescar lista"),
        ]),
        edit_info_label,
        ft.Container(height=10),
        edit_stage_dd,
        edit_location_dd,
        edit_semana_dd,
        ft.Container(height=15),
        ft.FilledButton(
            "Guardar cambios",
            icon=ft.Icons.SAVE,
            on_click=on_guardar_edicion,
            style=ft.ButtonStyle(bgcolor=ft.Colors.ORANGE, color=ft.Colors.WHITE),
            ref=edit_save_btn_ref,
            disabled=True,
        ),
        ft.Divider(),
        ft.Text("Actualización automática", size=16, weight=ft.FontWeight.BOLD),
        ft.Text(
            "Avanza +1 semana a todos los lotes según la semana ISO actual.\n"
            "También actualiza la etapa automáticamente según la semana.",
            size=11,
            color=ft.Colors.GREY_600,
        ),
    ], spacing=10, scroll=ft.ScrollMode.AUTO)
    
    def etapa_por_semana(semana):
        """Determina la etapa según la semana del lote."""
        semana = int(semana)
        if 1 <= semana <= 4:
            return 'CLONADO'
        elif 5 <= semana <= 7:
            return 'VEG. TEMPRANO'
        elif 8 <= semana <= 9:
            return 'VEG. TARDIO'
        elif 10 <= semana <= 20:
            return 'FLORACIÓN'
        elif semana == 21:
            return 'SECADO'
        elif semana == 22:
            return 'PT'
        else:
            return ''
    
    def actualizar_semanas_etapas_auto(e=None):
        """Actualiza semanas y etapas de todos los lotes según semana ISO."""
        try:
            lotes = leer_csv()
            hoy = datetime.now()
            semana_iso_actual = hoy.isocalendar()[1]
            
            cambios = []
            
            for lote in lotes:
                try:
                    sem = int(lote.get('Semana', '0'))
                except:
                    continue
                
                # Leer la semana ISO de la última actualización
                ultima_act = lote.get('ÚltimaActualización', '')
                semana_iso_lote = None
                if ultima_act:
                    try:
                        semana_iso_lote = datetime.strptime(ultima_act, '%Y-%m-%d').isocalendar()[1]
                    except:
                        semana_iso_lote = None
                
                # Solo avanzar si la semana ISO actual es distinta a la última registrada
                if 1 <= sem < 22 and (semana_iso_lote is None or semana_iso_lote != semana_iso_actual):
                    nueva_sem = sem + 1
                    nueva_etapa = etapa_por_semana(nueva_sem)
                    etapa_ant = lote.get('Stage', '')
                    
                    cambios.append({
                        'id': lote.get('ID', ''),
                        'sem_ant': sem,
                        'sem_nueva': nueva_sem,
                        'etapa_ant': etapa_ant,
                        'etapa_nueva': nueva_etapa,
                    })
                    
                    lote['Semana'] = str(nueva_sem)
                    lote['Stage'] = nueva_etapa
                    lote['ÚltimaActualización'] = hoy.strftime('%Y-%m-%d')
            
            if not cambios:
                # Usar diálogo en lugar de snackbar para mayor visibilidad
                def cerrar_info(e):
                    dlg_info.open = False
                    page.update()
                
                dlg_info = ft.AlertDialog(
                    modal=True,
                    title=ft.Text("✅ Sin cambios"),
                    content=ft.Text("Todos los lotes ya están actualizados para esta semana."),
                    actions=[ft.TextButton("OK", on_click=cerrar_info)],
                )
                page.overlay.append(dlg_info)
                dlg_info.open = True
                page.update()
                return
            
            # Mostrar diálogo de confirmación
            cambios_text = "\n".join([
                f"{c['id']}: Sem {c['sem_ant']}→{c['sem_nueva']} | {c['etapa_ant']}→{c['etapa_nueva']}"
                for c in cambios[:10]  # Mostrar máximo 10
            ])
            if len(cambios) > 10:
                cambios_text += f"\n... y {len(cambios) - 10} más"
            
            def cerrar_dialogo(e):
                dialogo.open = False
                page.update()
            
            def confirmar_actualizacion(e):
                dialogo.open = False
                page.update()
                
                # Guardar y sincronizar
                guardar_csv(lotes)
                async def do_upload():
                    success, msg = await asyncio.to_thread(subir_csv_github)
                    if not success:
                        show_snackbar(f"⚠️ No sincronizado: {msg}", error=True)
                    else:
                        page.snack_bar = ft.SnackBar(ft.Text(f"✅ {len(cambios)} lotes actualizados"))
                        page.snack_bar.open = True
                    # Refrescar listas
                    refresh_edit_lotes_popup()
                    refresh_lotes_dropdown()
                    page.update()
                asyncio.create_task(do_upload())
            
            dialogo = ft.AlertDialog(
                modal=True,
                title=ft.Text(f"¿Actualizar {len(cambios)} lotes?"),
                content=ft.Container(
                    content=ft.Text(cambios_text, size=12),
                    height=200,
                    width=300,
                ),
                actions=[
                    ft.TextButton("Cancelar", on_click=cerrar_dialogo),
                    ft.TextButton(
                        "Actualizar", 
                        on_click=confirmar_actualizacion,
                        style=ft.ButtonStyle(color=ft.Colors.PURPLE),
                    ),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
            )
            page.overlay.append(dialogo)
            dialogo.open = True
            page.update()
        except Exception as ex:
            page.snack_bar = ft.SnackBar(ft.Text(f"Error: {str(ex)}"))
            page.snack_bar.open = True
            page.update()
    
    # Agregar botón de actualización al tab_editar
    tab_editar.controls.append(
        ft.FilledButton(
            "⏰ Actualizar semanas y etapas",
            icon=ft.Icons.UPDATE,
            on_click=actualizar_semanas_etapas_auto,
            style=ft.ButtonStyle(bgcolor=ft.Colors.PURPLE, color=ft.Colors.WHITE),
        )
    )
    
    # ========== TAB 6: CONFIGURACIÓN ==========
    config_repo_field = ft.TextField(
        label="Repositorio GitHub",
        hint_text="usuario/nombre-repo",
        value=GITHUB_REPO or "",
        width=300,
        prefix_icon=ft.Icons.FOLDER,
    )
    
    config_token_field = ft.TextField(
        label="Token de GitHub",
        hint_text="ghp_xxxxxxxxxxxx",
        value=GITHUB_TOKEN or "",
        width=300,
        password=True,
        can_reveal_password=True,
        prefix_icon=ft.Icons.KEY,
    )
    
    config_status = ft.Text("", size=12)
    
    # Campo para usuario actual
    config_user_field = ft.TextField(
        label="Usuario actual",
        hint_text="Tu nombre",
        value=CURRENT_USER or "",
        width=300,
        prefix_icon=ft.Icons.PERSON,
        capitalization=ft.TextCapitalization.WORDS,
    )
    
    def on_save_user(e):
        nombre = config_user_field.value.strip()
        if not nombre:
            config_status.value = "❌ Ingresa un nombre de usuario"
            config_status.color = ft.Colors.RED
            page.update()
            return
        
        nombre_normalizado = guardar_usuario(nombre)
        config_user_field.value = nombre_normalizado
        config_status.value = f"✅ Usuario guardado: {nombre_normalizado}"
        config_status.color = ft.Colors.GREEN
        page.update()
    
    def on_save_config(e):
        repo = config_repo_field.value.strip()
        token = config_token_field.value.strip()
        user_input = config_user_field.value.strip()
        
        if not repo or "/" not in repo:
            config_status.value = "❌ Formato de repo inválido (usuario/repo)"
            config_status.color = ft.Colors.RED
            page.update()
            return
        
        if not token or len(token) < 10:
            config_status.value = "❌ Token inválido"
            config_status.color = ft.Colors.RED
            page.update()
            return
        
        # Determinar usuario a guardar: prioridad al campo de la UI, si no usar CURRENT_USER
        user_to_save = user_input if user_input else (CURRENT_USER if CURRENT_USER else None)
        if not user_to_save:
            config_status.value = "❌ Ingresa un usuario antes de guardar"
            config_status.color = ft.Colors.RED
            page.update()
            return
        
        async def guardar_async():
            # Llamar a la función que persiste la config (maneja Android vs Desktop)
            try:
                guardar_config_en_storage(page, repo, token, user=user_to_save)
            except Exception as e:
                print(f"Error iniciando guardado de config: {e}")

            # En Android, confirmar leyendo SharedPreferences para asegurar que se guardó
            if hasattr(sys, 'getandroidapilevel'):
                get_config_func = cargar_config_desde_storage(page)
                if get_config_func:
                    try:
                        await get_config_func()
                    except Exception as e:
                        print(f"Error leyendo SharedPreferences tras guardar: {e}")

            # Ahora comprobar que los valores están presentes (o en Desktop se escribieron ya)
            if GITHUB_TOKEN != token or GITHUB_REPO != repo:
                config_status.value = "❌ Error al guardar: sin token configurado"
                config_status.color = ft.Colors.RED
                page.update()
                return

            # Actualizar controles y estado
            config_repo_field.value = repo
            config_token_field.value = token
            config_user_field.value = user_to_save
            config_status.value = "✅ Configuración guardada"
            config_status.color = ft.Colors.GREEN
            # Actualizar estado de conexión inmediatamente según los nuevos valores
            try:
                check_and_update_connection_status()
            except Exception:
                pass

            # Si existía un sentinel, NO lo quitamos automáticamente al guardar la config.
            # La reactivación de subidas se realiza manualmente: completa la configuración y elimina el
            # sentinel '.no_auto_restore' fuera de la app si estás seguro.

            # Intentar descargar el CSV remoto automáticamente y refrescar la UI
            try:
                ok, msg = await asyncio.to_thread(descargar_csv_github)
                if ok:
                    try:
                        refresh_lotes_list_radios()
                    except Exception:
                        pass
                    try:
                        refresh_edit_lotes_popup()
                    except Exception:
                        pass
                    try:
                        refresh_lotes_list()
                    except Exception:
                        pass
                    try:
                        cnt = len(leer_csv())
                        show_snackbar(f"CSV descargado: {cnt} lotes cargados")
                    except Exception:
                        pass
                else:
                    # Mostrar mensaje no intrusivo pero informativo
                    try:
                        show_snackbar(f"No se pudo descargar CSV: {msg}", error=True)
                    except Exception:
                        pass
            except Exception as e:
                try:
                    print(f"[SYNC] Error descargando CSV tras guardar config: {e}")
                except Exception:
                    pass

            page.update()
        asyncio.create_task(guardar_async())
    
    def on_test_connection(e):
        config_status.value = "🔄 Probando conexión..."
        config_status.color = ft.Colors.BLUE
        page.update()

        async def do_test():
            # Ejecutar en hilo para no bloquear UI
            success, msg = await asyncio.to_thread(descargar_csv_github)
            if success:
                config_status.value = f"✅ Conexión exitosa"
                config_status.color = ft.Colors.GREEN
                update_status(True, "Conectado")
                # Refrescar UI y listas ahora que se descargó el CSV
                try:
                    refresh_lotes_list_radios()
                except Exception:
                    pass
                try:
                    refresh_edit_lotes_popup()
                except Exception:
                    pass
                try:
                    refresh_lotes_list()
                except Exception:
                    pass
                # Mostrar cuántos lotes se cargaron
                try:
                    count = len(leer_csv())
                    show_snackbar(f"Conexión OK, {count} lotes cargados")
                except Exception:
                    pass
            else:
                config_status.value = f"❌ Error: {msg}"
                config_status.color = ft.Colors.RED
                update_status(False, msg)
                # Si hubo conflicto, notificar al usuario
                if isinstance(msg, str) and 'Conflicto' in msg:
                    show_snackbar('Conflicto detectado: se guardaron backups. Revisa registros/', error=True)
            page.update()

        try:
            asyncio.create_task(do_test())
        except Exception:
            # Fallback síncrono si no hay loop
            success, msg = descargar_csv_github()
            if success:
                config_status.value = f"✅ Conexión exitosa"
                config_status.color = ft.Colors.GREEN
                update_status(True, "Conectado")
                try:
                    refresh_lotes_list_radios()
                except Exception:
                    pass
                try:
                    refresh_edit_lotes_popup()
                except Exception:
                    pass
                try:
                    refresh_lotes_list()
                except Exception:
                    pass
            else:
                config_status.value = f"❌ Error: {msg}"
                config_status.color = ft.Colors.RED
                update_status(False, msg)
            page.update()
    
    def on_clear_config(e):
        # Borrar config local (desktop)
        config_path = get_config_path()
        if os.path.exists(config_path):
            os.remove(config_path)
            
        # Borrar config persistente (Android)
        if hasattr(sys, 'getandroidapilevel'):
            try:
                from flet import SharedPreferences
                async def clear_prefs_and_globals():
                    try:
                        prefs = SharedPreferences()
                        await prefs.set("lotes_config", json.dumps({}))
                    except Exception as _:
                        # If set fails, try remove
                        try:
                            prefs = SharedPreferences()
                            await prefs.remove("lotes_config")
                        except Exception:
                            pass
                    # Ensure globals are cleared
                    try:
                        globals()["GITHUB_REPO"] = ""
                        globals()["GITHUB_TOKEN"] = ""
                        globals()["CURRENT_USER"] = ""
                    except Exception:
                        pass
                    try:
                        page.update()
                    except Exception:
                        pass
                try:
                    asyncio.create_task(clear_prefs_and_globals())
                except Exception:
                    try:
                        asyncio.run(clear_prefs_and_globals())
                    except Exception as err:
                        print(f"Error borrando SharedPreferences: {err}")
            except Exception as err:
                print(f"Error iniciando limpieza SharedPreferences: {err}")
        # Limpiar campos y estado en UI y memoria
        config_repo_field.value = ""
        config_token_field.value = ""
        config_user_field.value = ""
        # Vaciar usuario y credenciales en memoria también
        global CURRENT_USER, GITHUB_REPO, GITHUB_TOKEN, CONFIG_LAST_CLEARED
        CURRENT_USER = ""
        GITHUB_REPO = ""
        GITHUB_TOKEN = ""
        CONFIG_LAST_CLEARED = time.time()
        config_status.value = "🗑️ Configuración eliminada"
        config_status.color = ft.Colors.ORANGE
        # Reactivation UI removed; nothing to update here
        # Actualizar estado de conexión (mostrará que falta token/repo/usuario)
        try:
            check_and_update_connection_status()
        except Exception:
            pass

    def on_clear_local_data(e):
        """Borra los datos locales con backup y actualiza la UI inmediatamente."""
        def cerrar(e):
            dialog.open = False
            page.update()

        def confirmar(e):
            dialog.open = False
            page.update()
            # Crear backup del archivo original en registros pero NO modificar el archivo canonical LOTES_CSV
            b = None
            try:
                if os.path.exists(LOTES_CSV):
                    ensure_registros_dir()
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    dest = os.path.join(REGISTROS_DIR, f"lotes_template_deleted_{timestamp}.csv")
                    shutil.copy2(LOTES_CSV, dest)
                    b = dest
                    show_snackbar(f'Archivo original preservado en: {os.path.basename(dest)}')
                else:
                    # Si no existe el archivo canonical, aún creamos registros de estado
                    ensure_registros_dir()
            except Exception as ex:
                print(f"Error respaldando LOTES_CSV: {ex}")

            # Si existe un archivo de trabajo previo, eliminarlo para mostrar 'limpio' en UI
            try:
                if os.path.exists(LOTES_WORKING):
                    try:
                        os.remove(LOTES_WORKING)
                    except Exception:
                        pass
            except Exception:
                pass

            # Marcar estado: NO tocar LOTES_CSV, pero indicar que local está 'borrado' (ui vacía)
            try:
                with open(NO_AUTO_RESTORE_FILE, 'w', encoding='utf-8') as f:
                    f.write(datetime.now().isoformat())
                globals()["LOCAL_DATA_CLEARED"] = True
                show_snackbar('Datos locales marcados como borrados; archivo original preservado')
            except Exception:
                pass

            # Limpiar meta (eliminar o reiniciar)
            try:
                meta_path = get_local_meta_path()
                if os.path.exists(meta_path):
                    try:
                        os.remove(meta_path)
                    except Exception:
                        # Fallback a sobreescribir
                        save_local_meta({})
                else:
                    save_local_meta({})
            except Exception:
                pass

            # Refrescar listas y UI (y limpiar controles redundantes)
            try:
                refresh_lotes_list()
                refresh_edit_lotes_popup()
                try:
                    refresh_lotes_list_radios()
                except Exception:
                    pass
                # Asegurarse de limpiar controles de edición que puedan quedar con datos
                try:
                    edit_lote_popup.items.clear()
                    edit_lote_selector_text.value = "No hay lotes"
                    edit_info_label.value = "No hay lotes para editar"
                    try:
                        edit_stage_dd.value = ""
                        edit_location_dd.value = ""
                        edit_semana_dd.value = ""
                        # Mantener vacíos pero habilitados para permitir nuevas selecciones
                        edit_stage_dd.disabled = False
                        edit_location_dd.disabled = False
                        edit_semana_dd.disabled = False
                        try:
                            edit_stage_dd.update()
                        except Exception:
                            pass
                        try:
                            edit_location_dd.update()
                        except Exception:
                            pass
                        try:
                            edit_semana_dd.update()
                        except Exception:
                            pass
                    except Exception:
                        pass
                    try:
                        if edit_save_btn_ref and edit_save_btn_ref.current:
                            edit_save_btn_ref.current.disabled = True
                            edit_save_btn_ref.current.update()
                    except Exception:
                        pass
                except Exception:
                    pass
            except Exception:
                pass
            # Mostrar resultado
            if b:
                show_snackbar(f"Backup guardado: {os.path.basename(b)}")
            if True:
                show_snackbar('Datos locales marcados como borrados (archivo original preservado)', error=False)
            else:
                show_snackbar('No había datos locales para borrar', error=True)
            try:
                update_status(False, 'Sin datos locales')
            except Exception:
                pass
            page.update()

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text('Borrar datos locales'),
            content=ft.Text('Se creará un backup del archivo original y se marcarán los datos locales como borrados (el archivo original no se modificará). ¿Deseas continuar?'),
            actions=[
                ft.TextButton('Cancelar', on_click=cerrar),
                ft.TextButton('Marcar como borrados', on_click=confirmar, style=ft.ButtonStyle(color=ft.Colors.RED)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        page.overlay.append(dialog)
        dialog.open = True
        page.update()
        # (visualización de config persistente eliminada)
        page.update()
    


    # Reactivation button removed: reactivation must be done by configuring GitHub and
    # manually removing the sentinel '.no_auto_restore' outside the app for safety.

    tab_config = ft.Column([
        ft.Text("👤 Usuario", size=20, weight=ft.FontWeight.BOLD),
        ft.Text(
            "Tu nombre aparecerá en los commits de GitHub.",
            size=12,
            color=ft.Colors.GREY_700,
        ),
        ft.Column([
            config_user_field,
            ft.Row([
                ft.IconButton(
                    icon=ft.Icons.SAVE,
                    icon_color=ft.Colors.GREEN,
                    tooltip="Guardar usuario",
                    on_click=on_save_user,
                ),
            ], alignment=ft.MainAxisAlignment.START),
        ]),
        ft.Divider(),
        ft.Text("⚙️ Configuración GitHub", size=20, weight=ft.FontWeight.BOLD),
        ft.Text(
            "Configura tu repositorio de GitHub para sincronizar los datos.",
            size=12,
            color=ft.Colors.GREY_700,
        ),
        ft.Container(height=10),
        config_repo_field,
        config_token_field,
        ft.Container(height=10),
        ft.Row([
            ft.FilledButton(
                "Guardar",
                icon=ft.Icons.SAVE,
                on_click=on_save_config,
                style=ft.ButtonStyle(bgcolor=ft.Colors.GREEN, color=ft.Colors.WHITE),
            ),
            ft.OutlinedButton(
                "Probar conexión",
                icon=ft.Icons.WIFI,
                on_click=on_test_connection,
            ),
            ft.TextButton(
                "Limpiar",
                icon=ft.Icons.DELETE_OUTLINE,
                on_click=on_clear_config,
            ),
        ], wrap=True),
        # Botón de borrar datos locales en su propia línea para mejor UX en móviles
        ft.Row([
            ft.TextButton(
                "Borrar datos locales",
                icon=ft.Icons.DELETE_FOREVER,
                on_click=on_clear_local_data,
                style=ft.ButtonStyle(color=ft.Colors.RED),
            ),
        ], alignment=ft.MainAxisAlignment.END),
        # Nota sobre sentinel y comportamiento seguro (se muestra cuando hay datos marcados como borrados)
        ft.Row([
            ft.Text(
                "⚠️ Si marcaste los datos como borrados, el archivo de trabajo local y el sentinel '.no_auto_restore' impiden subidas automáticas. No borramos 'lotes_template.csv'.",
                size=12,
                color=ft.Colors.GREY_700,
            ),
        ]),
        config_status,
    ], spacing=10, scroll=ft.ScrollMode.AUTO)

    # show_restore_remote_dialog removed: restoring remote from backup is disabled in the UI by design. Use external tools or manual GitHub restore if necessary.
    
    # ========== NAVEGACIÓN CON CONTENIDO ==========
    content_area = ft.Container(
        content=ft.Container(tab_crear, padding=15),
        expand=True,
    )
    
    def change_view(e):
        index = e.control.selected_index
        views = [tab_crear, tab_variedades, tab_editar, tab_graficos, tab_listado, tab_config]

        # Primero cambiar vista (aseguramos que los controles estén añadidos)
        content_area.content = ft.Container(views[index], padding=15)
        page.update()

        # Si vamos a la pestaña de Variedades, actualizar opciones y valor por defecto
        if index == 1:
            try:
                options = [ft.dropdown.Option(v) for v in sorted(VARIETIES)]
                variety_dd.options.clear()
                for opt in options:
                    variety_dd.options.append(opt)
                if options:
                    variety_dd.value = options[0].value
                try:
                    variety_dd.update()
                except Exception:
                    pass
            except Exception:
                pass

        # Si vamos a la pestaña de Config, asegúrese de actualizar los campos ahora que están añadidos
        if index == len(views) - 1:
            # Actualizar campos con valores globales
            config_repo_field.value = GITHUB_REPO or ""
            config_token_field.value = GITHUB_TOKEN or ""
            config_user_field.value = CURRENT_USER or ""
            # Intentar actualizar visualmente, pero sin lanzar excepción si aún no hay page
            try:
                config_repo_field.update()
            except Exception:
                pass
            try:
                config_token_field.update()
            except Exception:
                pass
            try:
                config_user_field.update()
            except Exception:
                pass
            # Si estamos en Android, intentar cargar desde SharedPreferences si los campos están vacíos
            if hasattr(sys, 'getandroidapilevel'):
                async def load_config_android_and_update():
                    # Evitar una carga inmediata desde SharedPreferences si acabamos de borrar la config
                    try:
                        if CONFIG_LAST_CLEARED and (time.time() - CONFIG_LAST_CLEARED) < 2.0:
                            # Forzar campos vacíos y actualizar UI
                            config_repo_field.value = ""
                            config_token_field.value = ""
                            config_user_field.value = ""
                            try:
                                config_repo_field.update()
                            except Exception:
                                pass
                            try:
                                config_token_field.update()
                            except Exception:
                                pass
                            try:
                                config_user_field.update()
                            except Exception:
                                pass
                    except Exception:
                        pass

                    get_config = cargar_config_desde_storage(page)
                    if get_config:
                        try:
                            await get_config()
                        except Exception:
                            pass

                    # Actualizar estado de conexión y refrescar UI
                    try:
                        check_and_update_connection_status()
                    except Exception:
                        pass
                    try:
                        page.update()
                    except Exception:
                        pass
                    return

                # Agendar la tarea de carga en Android
                try:
                    asyncio.create_task(load_config_android_and_update())
                except Exception:
                    pass

            # Reactivation UI removed: visibility update not needed
    
    nav_bar = ft.NavigationBar(
        selected_index=0,
        on_change=change_view,
        destinations=[
            ft.NavigationBarDestination(icon=ft.Icons.ADD_BOX, label="Crear"),
            ft.NavigationBarDestination(icon=ft.Icons.GRASS, label="Lotes"),
            ft.NavigationBarDestination(icon=ft.Icons.EDIT, label="Editar"),
            ft.NavigationBarDestination(icon=ft.Icons.ANALYTICS, label="Gráficos"),
            ft.NavigationBarDestination(icon=ft.Icons.LIST, label="Listado"),
            ft.NavigationBarDestination(icon=ft.Icons.SETTINGS, label="Config"),
        ],
    )
    
    # Layout principal (sync controls moved to bottom, aligned right without stretching)
    page.add(
        ft.Column([
            status_bar,
            content_area,
            ft.Row([ft.Container(expand=True), sync_controls], alignment=ft.MainAxisAlignment.END),
        ], expand=True),
    )
    page.navigation_bar = nav_bar
    
    # Inicialización (configuración asíncrona ya lanzada en on_load)
    # startup_restore y refresco de listas se pueden lanzar aquí si necesario
    
    # Refrescar listas (init_config se lanza desde on_page_load)
    refresh_lotes_list()
    refresh_edit_lotes_popup()
    # También poblar popup de selección de lotes
    try:
        refresh_lotes_list_radios()
    except Exception:
        pass

    # Lanzar init_config aquí: UI ya está añadida y los controles existen
    try:
        asyncio.create_task(init_config())
    except Exception as e:
        print(f"No se pudo lanzar init_config desde final de main: {e}")
    # (El diálogo de usuario solo se muestra si no hay usuario tras cargar config, ver on_page_load)


# Punto de entrada
if __name__ == "__main__":
    ft.app(main)  # Compatible con versiones anteriores también
